from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash
from datetime import datetime, timedelta
from collections import defaultdict, OrderedDict
import os
from models import db, User, Category, Product, Supplier, Branch, Customer, Purchase, PurchaseItem, Sale, SaleItem, Inventory, StorageLocation, ProductLocationStock, InventoryTransaction, Setting, ReferenceProduct, ReferenceCategory, Employee, StockIssue, DamageRecord, StockIssueRequest, EmployeeReturn, Stocktake, StocktakeItem, AuditLog
from forms import LoginForm, RegisterForm, CategoryForm, ProductForm, SupplierForm, CustomerForm, UserForm, SaleForm, CatalogImportForm, CategoryCatalogImportForm, BranchForm, StorageLocationForm, ProductLocationStockForm, EmployeeForm, StockIssueForm, DamageRecordForm, StockIssueRequestForm, EmployeeReturnForm, StocktakeForm, StocktakeItemForm
from sqlalchemy.orm import joinedload
from functools import wraps
from flask import redirect, url_for, flash
from werkzeug.utils import secure_filename
from flask_migrate import Migrate
import random
import string
import smtplib
import time
from io import BytesIO
from email.mime.text import MIMEText
from sqlalchemy import text, func
from sqlalchemy.exc import OperationalError
import json

PRIVILEGED_ROLES = ('admin', 'warehouse_manager')
APPROVAL_ROLES = ('admin', 'approver', 'warehouse_manager')
AUDIT_ROLES = ('admin', 'auditor', 'warehouse_manager')

PERMISSION_GROUPS = OrderedDict([
    ('الصفحات الرئيسية', [
        ('dashboard.view', 'عرض لوحة المتابعة'),
        ('products.view', 'عرض الأصناف'),
        ('products.manage', 'إضافة/تعديل/حذف الأصناف'),
        ('categories.view', 'عرض التصنيفات'),
        ('categories.manage', 'إضافة/تعديل/حذف التصنيفات'),
        ('suppliers.view', 'عرض الموردين'),
        ('suppliers.manage', 'إضافة/تعديل/حذف الموردين'),
        ('purchases.view', 'عرض التوريدات'),
        ('purchases.manage', 'إضافة واستلام التوريدات'),
        ('inventory.view', 'عرض المخزون والحركات'),
        ('inventory.adjust', 'تعديل المخزون'),
    ]),
    ('الفروع والعهد', [
        ('branches.view', 'عرض الفروع'),
        ('branches.manage', 'إضافة/تعديل الفروع'),
        ('employees.view', 'عرض الموظفين'),
        ('employees.manage', 'إضافة/تعديل الموظفين'),
        ('locations.view', 'عرض مواقع التخزين'),
        ('locations.manage', 'إضافة مواقع التخزين'),
        ('location_stock.manage', 'تعديل كميات مواقع التخزين'),
        ('issue_requests.view', 'عرض طلبات الصرف'),
        ('issue_requests.create', 'إنشاء طلب صرف'),
        ('issue_requests.approve', 'اعتماد/رفض طلبات الصرف'),
        ('issue_requests.execute', 'تنفيذ طلبات الصرف'),
        ('stock_issue.direct', 'صرف عهدة مباشر'),
        ('returns.manage', 'تسجيل مرتجع عهدة'),
        ('damage.manage', 'تسجيل الهالك'),
        ('stocktakes.view', 'عرض الجرد'),
        ('stocktakes.manage', 'إنشاء وإدخال الجرد'),
        ('stocktakes.approve', 'اعتماد/رفض الجرد'),
    ]),
    ('التقارير والإدارة', [
        ('reports.view', 'عرض التقارير'),
        ('reports.export', 'تصدير التقارير Excel/PDF'),
        ('activity.view', 'عرض سجل النشاط'),
        ('audit.view', 'عرض سجل التدقيق'),
        ('users.manage', 'إدارة المستخدمين والصلاحيات'),
        ('settings.manage', 'الإعدادات واستيراد الكتالوجات'),
    ]),
])

ALL_PERMISSION_CODES = tuple(code for items in PERMISSION_GROUPS.values() for code, _ in items)

ROLE_DEFAULT_PERMISSIONS = {
    'admin': set(ALL_PERMISSION_CODES),
    'warehouse_manager': {
        'dashboard.view', 'products.view', 'products.manage', 'categories.view', 'categories.manage',
        'suppliers.view', 'suppliers.manage', 'purchases.view', 'purchases.manage',
        'inventory.view', 'inventory.adjust', 'branches.view', 'branches.manage',
        'employees.view', 'employees.manage', 'locations.view', 'locations.manage',
        'location_stock.manage', 'issue_requests.view', 'issue_requests.create',
        'issue_requests.approve', 'issue_requests.execute', 'stock_issue.direct',
        'returns.manage', 'damage.manage', 'stocktakes.view', 'stocktakes.manage',
        'stocktakes.approve', 'reports.view', 'reports.export', 'activity.view', 'audit.view',
    },
    'approver': {'dashboard.view', 'issue_requests.view', 'issue_requests.approve'},
    'auditor': {'dashboard.view', 'inventory.view', 'issue_requests.view', 'stocktakes.view', 'stocktakes.manage', 'reports.view', 'reports.export', 'activity.view', 'audit.view'},
    'requester': {'dashboard.view', 'products.view', 'inventory.view', 'issue_requests.view', 'issue_requests.create'},
    'user': {'dashboard.view', 'products.view', 'inventory.view', 'issue_requests.view', 'issue_requests.create'},
    'cashier': {'dashboard.view'},
}

ENDPOINT_PERMISSIONS = {
    'index': 'dashboard.view',
    'dashboard': 'dashboard.view',
    'products': 'products.view',
    'add_product': 'products.manage',
    'edit_product': 'products.manage',
    'delete_product': 'products.manage',
    'categories': 'categories.view',
    'add_category': 'categories.manage',
    'edit_category': 'categories.manage',
    'delete_category': 'categories.manage',
    'suppliers': 'suppliers.view',
    'add_supplier': 'suppliers.manage',
    'edit_supplier': 'suppliers.manage',
    'delete_supplier': 'suppliers.manage',
    'purchases': 'purchases.view',
    'add_purchase': 'purchases.manage',
    'receive_purchase': 'purchases.manage',
    'purchases_due': 'purchases.manage',
    'notify_due_purchase': 'purchases.manage',
    'mark_purchase_paid': 'purchases.manage',
    'inventory': 'inventory.view',
    'inventory_adjustment': 'inventory.adjust',
    'product_transactions': 'inventory.view',
    'branches': 'branches.view',
    'add_branch': 'branches.manage',
    'edit_branch': 'branches.manage',
    'employees': 'employees.view',
    'add_employee': 'employees.manage',
    'edit_employee': 'employees.manage',
    'storage_locations': 'locations.view',
    'add_storage_location': 'locations.manage',
    'storage_stocks': 'location_stock.manage',
    'issue_requests': 'issue_requests.view',
    'approve_issue_request': 'issue_requests.approve',
    'reject_issue_request': 'issue_requests.approve',
    'execute_issue_request': 'issue_requests.execute',
    'stock_issue': 'stock_issue.direct',
    'employee_returns': 'returns.manage',
    'damage_record': 'damage.manage',
    'stocktakes': 'stocktakes.view',
    'view_stocktake': 'stocktakes.view',
    'submit_stocktake': 'stocktakes.manage',
    'reject_stocktake': 'stocktakes.approve',
    'approve_stocktake': 'stocktakes.approve',
    'reports': 'reports.view',
    'purchases_report': 'reports.view',
    'inventory_report': 'reports.view',
    'issues_report': 'reports.view',
    'damage_report': 'reports.view',
    'top_selling_report': 'reports.view',
    'profit_report': 'reports.view',
    'customers_report': 'reports.view',
    'export_inventory_excel': 'reports.export',
    'export_inventory_pdf': 'reports.export',
    'export_issues_excel': 'reports.export',
    'export_issues_pdf': 'reports.export',
    'export_damage_excel': 'reports.export',
    'export_damage_pdf': 'reports.export',
    'activity': 'activity.view',
    'audit': 'audit.view',
    'users': 'users.manage',
    'add_user': 'users.manage',
    'edit_user': 'users.manage',
    'delete_user': 'users.manage',
    'settings': 'settings.manage',
    'catalog_import': 'settings.manage',
    'clear_reference_catalog': 'settings.manage',
    'category_catalog_import': 'settings.manage',
    'clear_reference_categories': 'settings.manage',
}


def parse_permissions(value):
    if not value:
        return set()
    try:
        data = json.loads(value)
        if isinstance(data, list):
            return {item for item in data if item in ALL_PERMISSION_CODES}
    except (TypeError, ValueError):
        return {item.strip() for item in value.split(',') if item.strip() in ALL_PERMISSION_CODES}
    return set()


def user_permissions(user):
    if not user or not user.is_authenticated:
        return set()
    if user.role == 'admin':
        return set(ALL_PERMISSION_CODES)
    permissions = set(ROLE_DEFAULT_PERMISSIONS.get(user.role, set()))
    permissions.update(parse_permissions(getattr(user, 'custom_permissions', None)))
    return permissions


def has_permission(permission_code):
    if not permission_code:
        return current_user.is_authenticated
    return permission_code in user_permissions(current_user)


def permission_required(permission_code):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or not has_permission(permission_code):
                flash('ليس لديك صلاحية تنفيذ هذا الإجراء.', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        from flask_login import current_user
        if not current_user.is_authenticated or not has_permission(ENDPOINT_PERMISSIONS.get(request.endpoint, 'users.manage')):
            flash('ليس لديك صلاحية الوصول لهذه الصفحة.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            permission_code = ENDPOINT_PERMISSIONS.get(request.endpoint)
            if not current_user.is_authenticated or (current_user.role not in roles and not has_permission(permission_code)):
                flash('ليس لديك صلاحية تنفيذ هذا الإجراء.', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def warehouse_required(f):
    return roles_required('admin', 'warehouse_manager')(f)


def report_required(f):
    return roles_required('admin', 'warehouse_manager', 'auditor')(f)


def cashier_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        from flask_login import current_user
        if not current_user.is_authenticated or (current_user.role not in ('cashier', 'admin') and not has_permission(ENDPOINT_PERMISSIONS.get(request.endpoint))):
            flash('Access denied: cashier tab is required.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'
os.makedirs(app.instance_path, exist_ok=True)
database_path = os.path.join(app.instance_path, 'makhzan.db')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + database_path.replace('\\', '/')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'connect_args': {'timeout': 5},
    'pool_pre_ping': True
}

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
migrate = Migrate(app, db)


@app.before_request
def enforce_endpoint_permissions():
    if request.endpoint in (None, 'static', 'login', 'logout'):
        return None
    permission_code = ENDPOINT_PERMISSIONS.get(request.endpoint)
    if permission_code and current_user.is_authenticated and not has_permission(permission_code):
        flash('ليس لديك صلاحية الوصول لهذه الصفحة أو تنفيذ هذا الإجراء.', 'danger')
        return redirect(url_for('index'))
    return None


def is_database_locked_error(exc):
    return 'database is locked' in str(exc).lower()


def configure_sqlite():
    with app.app_context():
        with db.engine.connect() as conn:
            conn.exec_driver_sql('PRAGMA busy_timeout=5000')


def commit_with_retry(attempts=3, delay=0.2):
    for attempt in range(attempts):
        try:
            db.session.commit()
            return True
        except OperationalError as exc:
            db.session.rollback()
            if not is_database_locked_error(exc) or attempt == attempts - 1:
                raise
            time.sleep(delay * (attempt + 1))
    return False


def audit_log(action, entity_type=None, entity_id=None, details=None):
    user_id = current_user.id if current_user and current_user.is_authenticated else None
    ip_address = request.remote_addr if request else None
    db.session.add(AuditLog(
        user_id=user_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        ip_address=ip_address
    ))


def generate_sequence(model, field_name, prefix):
    today = datetime.now().strftime('%y%m%d')
    base = f'{prefix}{today}'
    field = getattr(model, field_name)
    count = model.query.filter(field.like(f'{base}%')).count() + 1
    while True:
        value = f'{base}{count:04d}'
        if not model.query.filter(field == value).first():
            return value
        count += 1


def populate_user_form_choices(form):
    branches = Branch.query.filter_by(is_active=True).order_by(Branch.name).all()
    form.branch_id.choices = [(0, 'كل الفروع')] + [(branch.id, f'{branch.name} ({branch.code})') for branch in branches]
    form.permissions.choices = [(code, f'{label} [{code}]') for items in PERMISSION_GROUPS.values() for code, label in items]


def scoped_branch_id():
    if not current_user.is_authenticated:
        return None
    if current_user.role == 'admin':
        return None
    return getattr(current_user, 'branch_id', None)


def can_access_branch(branch_id):
    scope_id = scoped_branch_id()
    return not scope_id or not branch_id or int(branch_id) == int(scope_id)


def can_access_employee(employee):
    return employee is not None and can_access_branch(employee.branch_id)


def require_e_signature(action_label):
    password = request.form.get('signature_password', '')
    if not password or not current_user.check_password(password):
        flash(f'كلمة مرور التوقيع الإلكتروني غير صحيحة لتنفيذ: {action_label}.', 'danger')
        return None
    return f'{current_user.username} | {current_user.role} | {datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")} UTC | {request.remote_addr or ""}'


def can_approve_requests():
    return current_user.is_authenticated and has_permission('issue_requests.approve')


def can_manage_warehouse():
    return current_user.is_authenticated and has_permission('issue_requests.execute')


def excel_response(filename, headers, rows, title=None):
    try:
        from openpyxl import Workbook
    except ImportError:
        flash('تصدير Excel يحتاج مكتبة openpyxl.', 'danger')
        return redirect(request.referrer or url_for('reports'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'
    row_index = 1
    if title:
        ws.cell(row=row_index, column=1, value=title)
        row_index += 2
    ws.append(headers)
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def rtl_text(value):
    text_value = '' if value is None else str(value)
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display
        return get_display(arabic_reshaper.reshape(text_value))
    except Exception:
        return text_value


def pdf_response(filename, title, headers, rows, subtitle='', summary=None, landscape_mode=True):
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape, portrait
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        flash('تصدير PDF يحتاج مكتبة reportlab.', 'danger')
        return redirect(request.referrer or url_for('reports'))

    fonts_dir = os.path.join(os.environ.get('WINDIR', r'C:\Windows'), 'Fonts')
    font_path = os.path.join(fonts_dir, 'arial.ttf')
    bold_font_path = os.path.join(fonts_dir, 'arialbd.ttf')
    try:
        if 'ArabicReport' not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont('ArabicReport', font_path))
        if 'ArabicReportBold' not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont('ArabicReportBold', bold_font_path))
    except Exception:
        pass

    page_size = landscape(A4) if landscape_mode else portrait(A4)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=page_size, rightMargin=8 * mm, leftMargin=8 * mm, topMargin=8 * mm, bottomMargin=8 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ArabicTitle', parent=styles['Title'], fontName='ArabicReportBold', fontSize=14, alignment=TA_CENTER, leading=18, spaceAfter=4)
    subtitle_style = ParagraphStyle('ArabicSubtitle', parent=styles['Normal'], fontName='ArabicReport', fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor('#5b6472'), leading=12)
    cell_style = ParagraphStyle('ArabicCell', parent=styles['Normal'], fontName='ArabicReport', fontSize=7, alignment=TA_RIGHT, leading=9)
    header_style = ParagraphStyle('ArabicHeader', parent=cell_style, fontName='ArabicReportBold', alignment=TA_CENTER, textColor=colors.white)

    story = [
        Paragraph(rtl_text(title), title_style),
        Paragraph(rtl_text(subtitle or f'تاريخ التقرير: {datetime.now().strftime("%Y-%m-%d %H:%M")}'), subtitle_style),
        Spacer(1, 5 * mm)
    ]
    if summary:
        summary_data = [[Paragraph(rtl_text(label), header_style), Paragraph(rtl_text(value), cell_style)] for label, value in summary]
        summary_table = Table(summary_data, colWidths=[35 * mm, 35 * mm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#1f4e79')),
            ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#c9d3df')),
            ('ROWBACKGROUNDS', (1, 0), (1, -1), [colors.white, colors.HexColor('#f7f9fb')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 4 * mm))

    table_data = [[Paragraph(rtl_text(h), header_style) for h in headers]]
    for row in rows:
        table_data.append([Paragraph(rtl_text(cell), cell_style) for cell in row])

    usable_width = page_size[0] - doc.leftMargin - doc.rightMargin
    col_width = usable_width / max(1, len(headers))
    table = Table(table_data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#b8c2cc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f7fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


@login_manager.user_loader
def load_user(user_id):
    user = User.query.get(int(user_id))
    if user and not user.is_active:
        return None
    return user

# Create database tables and admin user
def ensure_due_columns():
    from sqlalchemy.exc import OperationalError
    with app.app_context():
        conn = db.engine.connect()
        needs_commit = False

        def try_add(table, definition):
            nonlocal needs_commit
            try:
                conn.execute(text(f'ALTER TABLE {table} ADD COLUMN {definition}'))
                needs_commit = True
            except OperationalError:
                pass

        try_add('purchases', 'amount_paid FLOAT NOT NULL DEFAULT 0')
        try_add('purchases', 'due_date DATE')
        try_add('purchases', 'notification_email TEXT')
        try_add('purchases', 'due_reminder_sent BOOLEAN NOT NULL DEFAULT 0')
        try_add('sales', 'amount_paid FLOAT NOT NULL DEFAULT 0')
        try_add('sales', 'due_date DATE')
        try_add('sales', 'notification_email TEXT')
        try_add('sales', 'due_reminder_sent BOOLEAN NOT NULL DEFAULT 0')
        try_add('sales', 'cashier_id INTEGER')

        if needs_commit:
            conn.connection.commit()
        conn.close()


def ensure_reference_columns():
    from sqlalchemy.exc import OperationalError
    with app.app_context():
        conn = db.engine.connect()
        added = False

        def try_add(definition):
            nonlocal added
            try:
                conn.execute(text(f'ALTER TABLE reference_products ADD COLUMN {definition}'))
                added = True
            except OperationalError:
                pass

        try_add('notes TEXT')

        if added:
            conn.connection.commit()
        conn.close()


def ensure_company_columns():
    from sqlalchemy.exc import OperationalError
    with app.app_context():
        conn = db.engine.connect()
        added = False

        def try_add(table, definition):
            nonlocal added
            try:
                conn.execute(text(f'ALTER TABLE {table} ADD COLUMN {definition}'))
                added = True
            except OperationalError:
                pass

        try_add('employees', 'employee_code TEXT')
        try_add('employees', 'branch_id INTEGER')
        try_add('users', 'branch_id INTEGER')
        try_add('users', 'custom_permissions TEXT')
        try_add('stock_issue_requests', 'approved_signature TEXT')
        try_add('stock_issue_requests', 'executed_signature TEXT')
        try_add('stock_issue_requests', 'rejected_signature TEXT')
        try_add('employee_returns', 'signed_signature TEXT')
        try_add('stocktakes', 'rejected_by_id INTEGER')
        try_add('stocktakes', 'rejected_at DATETIME')
        try_add('stocktakes', 'approved_signature TEXT')
        try_add('stocktakes', 'rejected_signature TEXT')
        try_add('stocktakes', 'rejection_reason TEXT')

        if added:
            conn.connection.commit()
        conn.close()


def ensure_document_numbers():
    with app.app_context():
        changed = False
        for item in StockIssueRequest.query.filter((StockIssueRequest.request_number == None) | (StockIssueRequest.request_number == '')).all():
            item.request_number = generate_sequence(StockIssueRequest, 'request_number', 'REQ')
            changed = True
        for item in EmployeeReturn.query.filter((EmployeeReturn.return_number == None) | (EmployeeReturn.return_number == '')).all():
            item.return_number = generate_sequence(EmployeeReturn, 'return_number', 'RET')
            changed = True
        for item in Stocktake.query.filter((Stocktake.count_number == None) | (Stocktake.count_number == '')).all():
            item.count_number = generate_sequence(Stocktake, 'count_number', 'CNT')
            changed = True
        if changed:
            commit_with_retry()


def create_tables():
    for attempt in range(5):
        try:
            with app.app_context():
                configure_sqlite()
                db.create_all()
                ensure_due_columns()
                ensure_reference_columns()
                ensure_company_columns()
                ensure_document_numbers()
                # Create admin user if not exists
                if not User.query.filter_by(username='admin').first():
                    admin = User(
                        username='admin',
                        email='admin@example.com',
                        full_name='Admin User',
                        role='admin',
                        is_active=True
                    )
                    admin.set_password('admin123')
                    db.session.add(admin)
                    commit_with_retry()
            return
        except OperationalError as exc:
            db.session.rollback()
            if not is_database_locked_error(exc) or attempt == 4:
                raise
            time.sleep(0.5 * (attempt + 1))


create_tables()


RETAIL_ONLY_PREFIXES = (
    '/sales',
    '/cashier',
    '/customers',
)

RETAIL_ONLY_REPORT_PATHS = (
    '/reports/sales',
    '/reports/top-selling',
    '/reports/profit',
    '/reports/customers',
)


@app.before_request
def hide_retail_pages_in_company_mode():
    path = request.path.rstrip('/') or '/'
    if path.startswith('/static') or not current_user.is_authenticated:
        return None
    if any(path == prefix or path.startswith(f'{prefix}/') for prefix in RETAIL_ONLY_PREFIXES):
        flash('تم إخفاء صفحات البيع والعملاء والكاشير لأن النظام مضبوط كمخزن شركة داخلي.', 'warning')
        return redirect(url_for('index'))
    if path in RETAIL_ONLY_REPORT_PATHS:
        flash('هذا التقرير خاص بالبيع التجاري وغير مستخدم في نظام مخازن الشركة.', 'warning')
        return redirect(url_for('reports'))
    return None


class SaleProcessingError(Exception):
    """Raised when sale processing fails due to validation or inventory issues."""
    pass


def generate_invoice_number(prefix='S'):
    base = datetime.now().strftime('%y%m%d')
    while True:
        candidate = f'{prefix}{base}{random.randint(1000, 9999)}'
        if not Sale.query.filter_by(invoice_number=candidate).first():
            return candidate


def get_or_create_setting(key, default=''):
    setting = Setting.query.filter_by(key=key).first()
    if not setting:
        setting = Setting(key=key, value=default)
        db.session.add(setting)
    return setting
def create_sale_from_form(form, form_data, cashier_id):
    invoice_number = (form.invoice_number.data or '').strip()
    if invoice_number:
        if Sale.query.filter_by(invoice_number=invoice_number).first():
            raise SaleProcessingError('Invoice number already exists.')
    else:
        invoice_number = generate_invoice_number()
    form.invoice_number.data = invoice_number
    sale = Sale(
        invoice_number=invoice_number,
        customer_id=form.customer_id.data if form.customer_id.data not in (None, 0) else None,
        sale_date=form.sale_date.data,
        payment_method=form.payment_method.data,
        notes=form.notes.data,
        total_amount=0,
        amount_paid=0,
        due_date=None,
        notification_email=None,
        due_reminder_sent=False,
        cashier_id=cashier_id
    )
    db.session.add(sale)
    db.session.flush()
    total_amount = 0
    items_index = 0
    added_item = False
    while True:
        item_key = f'items-{items_index}-product_id'
        if item_key not in form_data:
                break
        product_id = form_data.get(item_key)
        try:
            product_id = int(product_id)
        except (TypeError, ValueError):
            items_index += 1
            continue
        quantity = form_data.get(f'items-{items_index}-quantity')
        price = form_data.get(f'items-{items_index}-price')
        try:
            quantity = int(quantity)
            price = float(price)
        except (TypeError, ValueError):
            items_index += 1
            continue
        if quantity <= 0 or price <= 0:
            items_index += 1
            continue
        inventory = Inventory.query.filter_by(product_id=product_id).first()
        product = Product.query.get(product_id)
        available = inventory.quantity if inventory else 0
        if not inventory or inventory.quantity < quantity:
            name = product.name if product else f'ID {product_id}'
            raise SaleProcessingError(f'Not enough stock for {name}. Available: {available}.')
        sale_item = SaleItem(
            sale_id=sale.id,
            product_id=product_id,
            quantity=quantity,
            price=price
        )
        db.session.add(sale_item)
        inventory.quantity -= quantity
        transaction = InventoryTransaction(
            product_id=product_id,
            quantity_before=inventory.quantity + quantity,
            quantity_change=-quantity,
            quantity_after=inventory.quantity,
            transaction_type='sale',
            reference_type='sale',
            reference_id=sale.id,
            notes=f'Sale - invoice {sale.invoice_number}',
            user_id=cashier_id
        )
        db.session.add(transaction)
        total_amount += quantity * price
        added_item = True
        items_index += 1
    if not added_item or total_amount <= 0:
        raise SaleProcessingError('Add at least one product with a valid price and quantity.')
    amount_paid, due_date, notification_email = extract_due_details(total_amount, form_data)
    sale.total_amount = total_amount
    sale.amount_paid = amount_paid
    sale.due_date = due_date
    sale.notification_email = notification_email
    sale.due_reminder_sent = False
    return sale



# Routes
@app.route('/')
@login_required
def index():
    total_products = Product.query.count()
    total_suppliers = Supplier.query.count()
    total_employees = Employee.query.filter_by(is_active=True).count()
    total_branches = Branch.query.filter_by(is_active=True).count()

    recent_purchases = Purchase.query.order_by(Purchase.created_at.desc()).limit(5).all()
    recent_issues = StockIssue.query.order_by(StockIssue.created_at.desc()).limit(5).all()
    recent_damage = DamageRecord.query.order_by(DamageRecord.created_at.desc()).limit(5).all()

    low_stock_products = db.session.query(Product, Inventory).join(
        Inventory, Product.id == Inventory.product_id
    ).filter(Inventory.quantity <= Product.min_quantity).limit(5).all()

    inventory_rows = Inventory.query.options(joinedload(Inventory.product)).all()
    total_inventory_quantity = sum(item.quantity for item in inventory_rows)
    total_inventory_value = sum(item.quantity * (item.product.purchase_price or 0) for item in inventory_rows if item.product)

    today = datetime.now().date()
    month_start = today.replace(day=1)
    month_issues = StockIssue.query.filter(StockIssue.issue_date >= month_start).all()
    month_damage = DamageRecord.query.filter(DamageRecord.damage_date >= month_start).all()
    month_issue_quantity = sum(issue.quantity for issue in month_issues)
    month_damage_quantity = sum(record.quantity for record in month_damage)

    due_purchases = Purchase.query.filter(
        Purchase.payment_method == 'majel',
        Purchase.status != 'cancelled',
        Purchase.amount_paid < Purchase.total_amount
    ).all()
    pending_due_purchase_total = sum(purchase.due_amount for purchase in due_purchases)

    return render_template(
        'index.html',
        total_products=total_products,
        total_suppliers=total_suppliers,
        total_employees=total_employees,
        total_branches=total_branches,
        recent_purchases=recent_purchases,
        recent_issues=recent_issues,
        recent_damage=recent_damage,
        low_stock_products=low_stock_products,
        total_inventory_quantity=total_inventory_quantity,
        total_inventory_value=total_inventory_value,
        month_issue_quantity=month_issue_quantity,
        month_damage_quantity=month_damage_quantity,
        pending_due_purchase_total=pending_due_purchase_total,
        pending_due_purchase_count=len(due_purchases)
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.is_active and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'danger')
    
    # Pass the current datetime to the template
    return render_template('login.html', form=form, now=datetime.now())

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# Categories routes
@app.route('/categories')
@login_required
def categories():
    categories = Category.query.all()
    form = CategoryForm()
    return render_template('categories/index.html', categories=categories, form=form)

@app.route('/categories/add', methods=['POST'])
@login_required
@warehouse_required
def add_category():
    form = CategoryForm()
    if form.validate_on_submit():
        category = Category(
            name=form.name.data,
            description=form.description.data
        )
        db.session.add(category)
        db.session.commit()
        flash('تم إضافة التصنيف بنجاح', 'success')
        return redirect(url_for('categories'))
    # إذا فشل الفاليديشن، أعرض نفس صفحة التصنيفات مع الأخطاء
    categories = Category.query.all()
    flash('حدث خطأ في البيانات المدخلة. يرجى التحقق.', 'danger')
    return render_template('categories/index.html', categories=categories, form=form)

@app.route('/categories/edit/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def edit_category(id):
    category = Category.query.get_or_404(id)
    
    # استخدام request.form بدلاً من form
    if request.method == 'POST':
        category.name = request.form.get('name')
        category.description = request.form.get('description')
        db.session.commit()
        flash('تم تحديث التصنيف بنجاح', 'success')
    
    return redirect(url_for('categories'))

@app.route('/categories/delete/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def delete_category(id):
    category = Category.query.get_or_404(id)
    db.session.delete(category)
    notification_sender = Setting.query.filter_by(key='notification_sender_email').first()
    if not notification_sender:
        notification_sender = Setting(key='notification_sender_email', value='')
        db.session.add(notification_sender)

    notification_password = Setting.query.filter_by(key='notification_password').first()
    if not notification_password:
        notification_password = Setting(key='notification_password', value='')
        db.session.add(notification_password)

    notification_sender_name = Setting.query.filter_by(key='notification_sender_name').first()
    if not notification_sender_name:
        notification_sender_name = Setting(key='notification_sender_name', value='')
        db.session.add(notification_sender_name)

    db.session.commit()
    flash('تم حذف التصنيف بنجاح', 'success')
    return redirect(url_for('categories'))

# Products routes
@app.route('/products')
@login_required
def products():
    products_page = Product.query.options(joinedload(Product.category), joinedload(Product.inventory)).order_by(Product.name).paginate(
        page=request.args.get('page', 1, type=int),
        per_page=50,
        error_out=False
    )
    return render_template('products/index.html', products=products_page.items, products_page=products_page)

@app.route('/products/add', methods=['GET', 'POST'])
@login_required
@warehouse_required
def add_product():
    form = ProductForm()
    
    # Get all categories for the dropdown
    categories = Category.query.all()
    
    # Check if there are any categories
    if not categories:
        # Create some default categories if none exist
        default_categories = [
            # Category(name='ألبان', description='منتجات الألبان والأجبان'),
            # Category(name='مشروبات', description='المشروبات الغازية والعصائر'),
            # Category(name='معلبات', description='الأطعمة المعلبة'),
            # Category(name='منظفات', description='منتجات التنظيف'),
            # Category(name='حلويات', description='الحلويات والشوكولاتة')
        ]
        db.session.add_all(default_categories)
        db.session.commit()
        categories = Category.query.all()
    
    # Set the choices for the category dropdown
    form.category_id.choices = [(c.id, c.name) for c in categories]
    form.category_id.choices.insert(0, (0, 'بدون تصنيف'))
    
    if form.validate_on_submit():
        # image_filename = None
        # if form.image.data:
        #     image_file = form.image.data
        #     image_filename = secure_filename(image_file.filename)
        #     upload_folder = os.path.join('static', 'uploads')
        #     os.makedirs(upload_folder, exist_ok=True)
        #     image_path = os.path.join(upload_folder, image_filename)
        #     image_file.save(image_path)
        # معالجة sku
        sku = form.sku.data.strip() if form.sku.data else ''
        if not sku:
            # توليد SKU عشوائي
            while True:
                sku = 'SKU-' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
                if not Product.query.filter_by(sku=sku).first():
                    break
        else:
# تحقق من عدم تكرار الـ SKU
            if Product.query.filter_by(sku=sku).first():
                flash('رمز المنتج (SKU) مستخدم من قبل، يرجى اختيار رمز آخر أو تركه فارغاً ليتم توليده تلقائياً.', 'danger')
                return render_template('products/add.html', form=form)
        selected_category_id = form.category_id.data if form.category_id.data not in (None, 0) else None
        custom_category_name = (form.category_autocomplete.data or '').strip()
        if not selected_category_id and custom_category_name:
            existing_category = Category.query.filter(func.lower(Category.name) == custom_category_name.lower()).first()
            if not existing_category:
                existing_category = Category(name=custom_category_name)
                db.session.add(existing_category)
                db.session.flush()
            selected_category_id = existing_category.id
        product = Product(
            name=form.name.data,
            description=form.description.data,
            barcode=form.barcode.data,
            sku=sku,
            purchase_price=form.purchase_price.data,
            sale_price=form.sale_price.data,
            min_quantity=form.min_quantity.data,
            category_id=selected_category_id
            # image=image_filename (تم الحذف)
        )
        db.session.add(product)
        db.session.commit()
        
        # Create inventory record
        inventory = Inventory(
            product_id=product.id,
            quantity=form.initial_quantity.data
        )
        db.session.add(inventory)
        
        # Create inventory transaction
        if form.initial_quantity.data > 0:
            transaction = InventoryTransaction(
                product_id=product.id,
                quantity_before=0,
                quantity_change=form.initial_quantity.data,
                quantity_after=form.initial_quantity.data,
                transaction_type='adjustment',
                reference_type='adjustment',
                notes='الكمية الأولية عند إضافة المنتج',
                user_id=current_user.id
            )
            db.session.add(transaction)
        
        db.session.commit()
        flash('تم إضافة المنتج بنجاح', 'success')
        return redirect(url_for('products'))
    
    return render_template('products/add.html', form=form)

@app.route('/products/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@warehouse_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    form.category_id.choices = [(c.id, c.name) for c in Category.query.all()]
    form.category_id.choices.insert(0, (0, 'بدون تصنيف'))
    
    if request.method == 'GET':
        form.category_id.data = product.category_id if product.category_id else 0
        if product.inventory:
            form.initial_quantity.data = product.inventory.quantity
    
    if form.validate_on_submit():
        old_quantity = product.inventory.quantity if product.inventory else 0
        
        product.name = form.name.data
        product.description = form.description.data
        product.barcode = form.barcode.data
        product.sku = form.sku.data
        product.purchase_price = form.purchase_price.data
        product.sale_price = form.sale_price.data
        product.min_quantity = form.min_quantity.data
        product.category_id = form.category_id.data if form.category_id.data != 0 else None
        
        # Update inventory
        if not product.inventory:
            inventory = Inventory(
                product_id=product.id,
                quantity=form.initial_quantity.data
            )
            db.session.add(inventory)
        else:
            if old_quantity != form.initial_quantity.data:
                # Create inventory transaction
                transaction = InventoryTransaction(
                    product_id=product.id,
                    quantity_before=old_quantity,
                    quantity_change=form.initial_quantity.data - old_quantity,
                    quantity_after=form.initial_quantity.data,
                    transaction_type='adjustment',
                    reference_type='adjustment',
                    notes='تعديل الكمية من خلال تحرير المنتج',
                    user_id=current_user.id
                )
                db.session.add(transaction)
                
                product.inventory.quantity = form.initial_quantity.data
        
        db.session.commit()
        flash('تم تحديث المنتج بنجاح', 'success')
        return redirect(url_for('products'))
    
    return render_template('products/edit.html', form=form, product=product)

@app.route('/products/delete/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def delete_product(id):
    product = Product.query.get_or_404(id)
    # احذف كل حركات المخزون المرتبطة بالمنتج
    InventoryTransaction.query.filter_by(product_id=product.id).delete()
    SaleItem.query.filter_by(product_id=product.id).delete()
    PurchaseItem.query.filter_by(product_id=product.id).delete()
    # احذف سجل المخزون المرتبط بالمنتج أولاً
    if product.inventory:
        db.session.delete(product.inventory)
    db.session.delete(product)
    db.session.commit()
    flash('تم حذف المنتج بنجاح', 'success')
    return redirect(url_for('products'))

# Suppliers routes
@app.route('/suppliers')
@login_required
def suppliers():
    suppliers = Supplier.query.all()
    return render_template('suppliers/index.html', suppliers=suppliers)

@app.route('/suppliers/add', methods=['GET', 'POST'])
@login_required
@warehouse_required
def add_supplier():
    form = SupplierForm()
    if form.validate_on_submit():
        supplier = Supplier(
            name=form.name.data,
            contact_person=form.contact_person.data,
            phone=form.phone.data,
            email=form.email.data,
            address=form.address.data,
            notes=form.notes.data
        )
        db.session.add(supplier)
        db.session.commit()
        flash('تم إضافة المورد بنجاح', 'success')
        return redirect(url_for('suppliers'))
    
    return render_template('suppliers/add.html', form=form)

@app.route('/suppliers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@warehouse_required
def edit_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    form = SupplierForm(obj=supplier)
    
    if form.validate_on_submit():
        supplier.name = form.name.data
        supplier.contact_person = form.contact_person.data
        supplier.phone = form.phone.data
        supplier.email = form.email.data
        supplier.address = form.address.data
        supplier.notes = form.notes.data
        db.session.commit()
        flash('تم تحديث المورد بنجاح', 'success')
        return redirect(url_for('suppliers'))
    
    return render_template('suppliers/edit.html', form=form, supplier=supplier)

@app.route('/suppliers/delete/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def delete_supplier(id):
    supplier = Supplier.query.get_or_404(id)
    db.session.delete(supplier)
    db.session.commit()
    flash('تم حذف المورد بنجاح', 'success')
    return redirect(url_for('suppliers'))

# Customers routes
@app.route('/customers')
@login_required
def customers():
    customers = Customer.query.all()
    return render_template('customers/index.html', customers=customers)

@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_customer():
    form = CustomerForm()
    if form.validate_on_submit():
        customer = Customer(
            name=form.name.data,
            phone=form.phone.data,
            email=form.email.data,
            address=form.address.data,
            notes=form.notes.data
        )
        db.session.add(customer)
        db.session.commit()
        flash('تم إضافة العميل بنجاح', 'success')
        return redirect(url_for('customers'))
    
    return render_template('customers/add.html', form=form)

@app.route('/customers/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    form = CustomerForm(obj=customer)
    
    if form.validate_on_submit():
        customer.name = form.name.data
        customer.phone = form.phone.data
        customer.email = form.email.data
        customer.address = form.address.data
        customer.notes = form.notes.data
        db.session.commit()
        flash('تم تحديث العميل بنجاح', 'success')
        return redirect(url_for('customers'))
    
    return render_template('customers/edit.html', form=form, customer=customer)

@app.route('/customers/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_customer(id):
    customer = Customer.query.get_or_404(id)
    db.session.delete(customer)
    db.session.commit()
    flash('تم حذف العميل بنجاح', 'success')
    return redirect(url_for('customers'))

# Purchases routes
@app.route('/purchases')
@login_required
def purchases():
    purchases_page = Purchase.query.options(joinedload(Purchase.supplier)).order_by(Purchase.created_at.desc()).paginate(
        page=request.args.get('page', 1, type=int),
        per_page=50,
        error_out=False
    )
    return render_template('purchases/index.html', purchases=purchases_page.items, purchases_page=purchases_page)

@app.route('/purchases/add', methods=['GET', 'POST'])
@login_required
@warehouse_required
def add_purchase():
    from forms import PurchaseForm
    form = PurchaseForm()
    form.supplier_id.choices = [(s.id, s.name) for s in Supplier.query.all()]
    
    # Get all products for the dropdown
    products = Product.query.all()
    
    if request.method == 'POST':
        print('FORM DATA:', dict(request.form))
        # Process the form submission
        if form.validate_on_submit():
            # Create new purchase - remove user_id parameter
            purchase = Purchase(
                invoice_number=form.invoice_number.data,
                purchase_date=form.purchase_date.data,
                supplier_id=form.supplier_id.data,
                total_amount=0,
                payment_method=request.form.get('payment_method', 'cash'),
                notes=request.form.get('notes', ''),
                amount_paid=0,
                due_date=None,
                notification_email=None,
                due_reminder_sent=False
            )
            db.session.add(purchase)
            db.session.flush()  # Get the purchase ID
            total_amount = 0
            
            # Process purchase items
            items_count = 0
            while f'items-{items_count}-product_id' in request.form:
                product_id = request.form.get(f'items-{items_count}-product_id')
                quantity = int(request.form.get(f'items-{items_count}-quantity', 0))
                price = float(request.form.get(f'items-{items_count}-price', 0))
                
                if product_id and quantity > 0 and price > 0:
                    # Add purchase item (بدون total)
                    item = PurchaseItem(
                        purchase_id=purchase.id,
                        product_id=product_id,
                        quantity=quantity,
                        price=price
                    )
                    db.session.add(item)
                    total_amount += quantity * price
                    
                    # Update inventory
                    inventory = Inventory.query.filter_by(product_id=product_id).first()
                    if inventory:
                        old_quantity = inventory.quantity
                        inventory.quantity += quantity
                        
                        # Create inventory transaction
                        transaction = InventoryTransaction(
                            product_id=product_id,
                            quantity_before=old_quantity,
                            quantity_change=quantity,
                            quantity_after=old_quantity + quantity,
                            transaction_type='purchase',
                            reference_type='purchase',
                            reference_id=purchase.id,
                            user_id=current_user.id
                        )
                        db.session.add(transaction)
                
                items_count += 1
            
            amount_paid, due_date, notification_email = extract_due_details(total_amount, request.form)
            purchase.total_amount = total_amount
            purchase.amount_paid = amount_paid
            purchase.due_date = due_date
            purchase.notification_email = notification_email
            purchase.due_reminder_sent = False

            db.session.commit()
            flash('تم إضافة فاتورة الشراء بنجاح', 'success')
            return redirect(url_for('purchases'))
    
    return render_template('purchases/add.html', form=form, products=products)

@app.route('/purchases/view/<int:id>')
@login_required
def view_purchase(id):
    purchase = Purchase.query.options(
        joinedload(Purchase.items).joinedload(PurchaseItem.product)
    ).get_or_404(id)
    return render_template('purchases/view.html', purchase=purchase)

# Sales routes
@app.route('/sales')
@login_required
def sales():
    sales = Sale.query.order_by(Sale.created_at.desc()).all()
    total_sales = sum(sale.total_amount for sale in sales if sale.status == 'completed')
    pending_due_total = sum(
        sale.due_amount for sale in sales
        if sale.payment_method == 'majel' and sale.status != 'cancelled' and sale.due_amount > 0
    )
    pending_due_count = sum(
        1 for sale in sales
        if sale.payment_method == 'majel' and sale.status != 'cancelled' and sale.due_amount > 0
    )
    return render_template(
        'sales/index.html',
        sales=sales,
        total_sales=total_sales,
        pending_due_total=pending_due_total,
        pending_due_count=pending_due_count
    )

@app.route('/sales/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_sale():
    form = SaleForm()
    customers = Customer.query.all()
    form.customer_id.choices = [(0, '-- Select Customer --')] + [(c.id, c.name) for c in customers]
    products = Product.query.all()
    inventories = {inv.product_id: inv.quantity for inv in Inventory.query.all()}
    if not form.invoice_number.data:
        form.invoice_number.data = generate_invoice_number()
    form.sale_date.data = form.sale_date.data or datetime.now().date()
    if form.validate_on_submit():
        try:
            create_sale_from_form(form, request.form, current_user.id)
            db.session.commit()
            flash('تم إضافة فاتورة المبيعات بنجاح', 'success')
            return redirect(url_for('sales'))
        except SaleProcessingError as exc:
            db.session.rollback()
            flash(str(exc), 'danger')
    return render_template('sales/add.html', form=form, products=products, inventories=inventories)


@app.route('/sales/view/<int:id>')
@login_required
def view_sale(id):
    sale = Sale.query.options(
        joinedload(Sale.items).joinedload(SaleItem.product)
    ).get_or_404(id)
    return render_template('sales/view.html', sale=sale)

@app.route('/cashier/pos')
@login_required
@cashier_required
def cashier_pos():
    form = SaleForm()
    customers = Customer.query.all()
    form.customer_id.choices = [(0, '-- Select Customer --')] + [(c.id, c.name) for c in customers]
    products = Product.query.all()
    inventories = {inv.product_id: inv.quantity for inv in Inventory.query.all()}
    if not form.invoice_number.data:
        form.invoice_number.data = generate_invoice_number()
    form.sale_date.data = form.sale_date.data or datetime.now().date()
    form.payment_method.data = form.payment_method.data or 'cash'
    return render_template('cashier/pos.html', form=form, products=products, inventories=inventories)

@app.route('/cashier/product-info')
@login_required
@cashier_required
def cashier_product_info():
    code = (request.args.get('code') or '').strip()
    if not code:
      return jsonify({'error': 'Code missing.'}), 400
    product = None
    if code.isdigit():
        product = Product.query.filter_by(id=int(code)).first()
    if not product:
        product = Product.query.filter((Product.barcode == code) | (Product.sku == code)).first()
    if not product:
      return jsonify({'error': 'Product not found.'}), 404
    inventory = Inventory.query.filter_by(product_id=product.id).first()
    return jsonify({
        'id': product.id,
        'name': product.name,
        'price': product.sale_price,
        'available': inventory.quantity if inventory else 0,
        'barcode': product.barcode or '',
        'sku': product.sku or ''
    })

@app.route('/cashier/checkout', methods=['POST'])
@login_required
@cashier_required
def cashier_checkout():
    form = SaleForm()
    customers = Customer.query.all()
    form.customer_id.choices = [(0, '-- Select Customer --')] + [(c.id, c.name) for c in customers]
    if form.validate_on_submit():
        try:
            create_sale_from_form(form, request.form, current_user.id)
        except SaleProcessingError as exc:
            db.session.rollback()
            flash(str(exc), 'danger')
            return redirect(url_for('cashier_pos'))
        db.session.commit()
        flash('تم تسجيل البيع عن طريق الكاشير.', 'success')
        return redirect(url_for('cashier_pos'))
    flash('Please complete the form correctly.', 'danger')
    return redirect(url_for('cashier_pos'))

@app.route('/cashier/logs')
@login_required
@admin_required
def cashier_logs():
    cash_sales = Sale.query.options(joinedload(Sale.cashier)).filter(Sale.cashier_id.isnot(None)).order_by(Sale.created_at.desc()).all()
    return render_template('cashier/logs.html', sales=cash_sales)

@app.route('/sales/cancel/<int:id>', methods=['POST'])
@login_required
@admin_required
def cancel_sale(id):
    sale = Sale.query.get_or_404(id)
    if sale.status == 'cancelled':
        flash('تم إلغاء هذه الفاتورة بالفعل!', 'warning')
        return redirect(url_for('sales'))
    sale.status = 'cancelled'
    for item in sale.items:
        inventory = Inventory.query.filter_by(product_id=item.product_id).first()
        if inventory:
            inventory.quantity += item.quantity
            transaction = InventoryTransaction(
                product_id=item.product_id,
                quantity_before=inventory.quantity - item.quantity,
                quantity_change=item.quantity,
                quantity_after=inventory.quantity,
                transaction_type='cancel_sale',
                reference_type='sale',
                reference_id=sale.id,
                notes=f'إلغاء فاتورة بيع رقم {sale.invoice_number}',
                user_id=current_user.id
            )
            db.session.add(transaction)
    db.session.commit()
    flash('تم إلغاء الفاتورة وإعادة المنتجات للمخزون بنجاح!', 'success')
    return redirect(url_for('sales'))

# Inventory routes
@app.route('/inventory')
@login_required
def inventory():
    status = request.args.get('status', 'all')
    
    query = db.session.query(Product, Inventory).join(
        Inventory, Product.id == Inventory.product_id
    )
    
    if status == 'low':
        query = query.filter(Inventory.quantity <= Product.min_quantity, Inventory.quantity > 0)
    elif status == 'out':
        query = query.filter(Inventory.quantity <= 0)
    
    inventory_page = query.order_by(Product.name).paginate(
        page=request.args.get('page', 1, type=int),
        per_page=50,
        error_out=False
    )
    inventory_items = inventory_page.items
    
    total_value = sum(item.quantity * item.product.purchase_price for _, item in inventory_items)
    
    return render_template('inventory/index.html', 
                          inventory=inventory_items, 
                          inventory_page=inventory_page,
                          total_value=total_value,
                          status=status)

@app.route('/inventory/adjustment', methods=['GET', 'POST'])
@login_required
@warehouse_required
def inventory_adjustment():
    from forms import InventoryAdjustmentForm
    form = InventoryAdjustmentForm()
    from models import Product, Inventory, InventoryTransaction
    products = Product.query.all()
    form.product_id.choices = [(p.id, p.name) for p in products]

    if form.validate_on_submit():
        product_id = form.product_id.data
        adjustment_type = form.adjustment_type.data
        quantity = form.quantity.data
        notes = form.notes.data
        inventory = Inventory.query.filter_by(product_id=product_id).first()
        if not inventory:
            flash('لم يتم العثور على مخزون لهذا المنتج.', 'danger')
            return render_template('inventory/adjustment.html', form=form)
        old_quantity = inventory.quantity
        if adjustment_type == 'add':
            inventory.quantity += quantity
            change = quantity
        else:
            if inventory.quantity < quantity:
                flash('الكمية المدخلة أكبر من الكمية المتوفرة في المخزون.', 'danger')
                return render_template('inventory/adjustment.html', form=form)
            inventory.quantity -= quantity
            change = -quantity
        # سجل حركة المخزون
        transaction = InventoryTransaction(
            product_id=product_id,
            quantity_before=old_quantity,
            quantity_change=change,
            quantity_after=inventory.quantity,
            transaction_type='adjustment',
            reference_type='adjustment',
            notes=notes,
            user_id=current_user.id
        )
        from models import db
        db.session.add(transaction)
        db.session.commit()
        flash('تم تعديل المخزون بنجاح.', 'success')
        return redirect(url_for('inventory'))

    return render_template('inventory/adjustment.html', form=form)

@app.route('/product/transactions/<int:id>')
@login_required
def product_transactions(id):
    product = Product.query.get_or_404(id)
    transactions = InventoryTransaction.query.filter_by(product_id=id).order_by(InventoryTransaction.timestamp.desc()).all()
    return render_template('inventory/transactions.html', product=product, transactions=transactions)


def populate_employee_form_choices(form):
    branches_query = Branch.query.filter_by(is_active=True)
    if scoped_branch_id():
        branches_query = branches_query.filter(Branch.id == scoped_branch_id())
    branches = branches_query.order_by(Branch.name).all()
    form.branch_id.choices = [(0, 'بدون فرع')] + [(branch.id, f'{branch.name} ({branch.code})') for branch in branches]


def populate_product_employee_choices(form):
    products = Product.query.order_by(Product.name).all()
    employees_query = Employee.query.options(joinedload(Employee.branch)).filter_by(is_active=True)
    if scoped_branch_id():
        employees_query = employees_query.filter(Employee.branch_id == scoped_branch_id())
    employees_list = employees_query.order_by(Employee.name).all()
    form.product_id.choices = [(p.id, p.name) for p in products]
    form.employee_id.choices = [
        (e.id, f'{e.employee_code or "-"} - {e.name} - {e.branch.name if e.branch else "بدون فرع"}')
        for e in employees_list
    ]


@app.route('/branches')
@login_required
@warehouse_required
def branches():
    branches_query = Branch.query
    if scoped_branch_id():
        branches_query = branches_query.filter(Branch.id == scoped_branch_id())
    branches = branches_query.order_by(Branch.is_active.desc(), Branch.name).all()
    form = BranchForm()
    return render_template('branches/index.html', branches=branches, form=form)


@app.route('/branches/add', methods=['POST'])
@login_required
@warehouse_required
def add_branch():
    if scoped_branch_id():
        flash('ليس لديك صلاحية إنشاء فرع جديد خارج نطاقك.', 'danger')
        return redirect(url_for('branches'))
    form = BranchForm()
    if form.validate_on_submit():
        code = (form.code.data or '').strip()
        if Branch.query.filter(func.lower(Branch.code) == code.lower()).first():
            flash('كود الفرع موجود بالفعل.', 'danger')
            return redirect(url_for('branches'))
        branch = Branch(
            name=form.name.data.strip(),
            code=code,
            location=form.location.data,
            notes=form.notes.data,
            is_active=form.is_active.data
        )
        db.session.add(branch)
        commit_with_retry()
        flash('تم إضافة الفرع بنجاح.', 'success')
    else:
        flash('راجع بيانات الفرع المدخلة.', 'danger')
    return redirect(url_for('branches'))


@app.route('/branches/edit/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def edit_branch(id):
    branch = Branch.query.get_or_404(id)
    if not can_access_branch(branch.id):
        flash('ليس لديك صلاحية تعديل هذا الفرع.', 'danger')
        return redirect(url_for('branches'))
    form = BranchForm()
    if form.validate_on_submit():
        code = (form.code.data or '').strip()
        duplicate = Branch.query.filter(func.lower(Branch.code) == code.lower(), Branch.id != id).first()
        if duplicate:
            flash('كود الفرع موجود بالفعل.', 'danger')
            return redirect(url_for('branches'))
        branch.name = form.name.data.strip()
        branch.code = code
        branch.location = form.location.data
        branch.notes = form.notes.data
        branch.is_active = form.is_active.data
        commit_with_retry()
        flash('تم تحديث بيانات الفرع.', 'success')
    else:
        flash('راجع بيانات الفرع المدخلة.', 'danger')
    return redirect(url_for('branches'))


def populate_storage_location_choices(form):
    branches_query = Branch.query.filter_by(is_active=True)
    if scoped_branch_id():
        branches_query = branches_query.filter(Branch.id == scoped_branch_id())
    branches = branches_query.order_by(Branch.name).all()
    form.branch_id.choices = [(0, 'بدون فرع')] + [(branch.id, f'{branch.name} ({branch.code})') for branch in branches]


def populate_location_stock_choices(form):
    form.product_id.choices = [(p.id, p.name) for p in Product.query.order_by(Product.name).all()]
    locations_query = StorageLocation.query.options(joinedload(StorageLocation.branch)).filter_by(is_active=True)
    if scoped_branch_id():
        locations_query = locations_query.filter(StorageLocation.branch_id == scoped_branch_id())
    form.location_id.choices = [
        (loc.id, f'{loc.code} - {loc.name} - {loc.branch.name if loc.branch else "بدون فرع"}')
        for loc in locations_query.order_by(StorageLocation.code).all()
    ]


@app.route('/storage-locations')
@login_required
@roles_required('admin', 'warehouse_manager')
def storage_locations():
    locations_query = StorageLocation.query.options(joinedload(StorageLocation.branch))
    if scoped_branch_id():
        locations_query = locations_query.filter(StorageLocation.branch_id == scoped_branch_id())
    locations = locations_query.order_by(StorageLocation.is_active.desc(), StorageLocation.name).all()
    form = StorageLocationForm()
    populate_storage_location_choices(form)
    return render_template('storage_locations/index.html', locations=locations, form=form)


@app.route('/storage-locations/add', methods=['POST'])
@login_required
@roles_required('admin', 'warehouse_manager')
def add_storage_location():
    form = StorageLocationForm()
    populate_storage_location_choices(form)
    if form.validate_on_submit():
        if not can_access_branch(form.branch_id.data):
            flash('ليس لديك صلاحية إضافة موقع تخزين لهذا الفرع.', 'danger')
            return redirect(url_for('storage_locations'))
        code = (form.code.data or '').strip()
        if StorageLocation.query.filter(func.lower(StorageLocation.code) == code.lower()).first():
            flash('كود موقع التخزين موجود بالفعل.', 'danger')
            return redirect(url_for('storage_locations'))
        location = StorageLocation(
            branch_id=form.branch_id.data if form.branch_id.data else None,
            name=form.name.data.strip(),
            code=code,
            location_type=form.location_type.data,
            notes=form.notes.data,
            is_active=form.is_active.data
        )
        db.session.add(location)
        audit_log('create_storage_location', 'StorageLocation', None, f'إنشاء موقع تخزين {code}')
        commit_with_retry()
        flash('تم إضافة موقع التخزين.', 'success')
    else:
        flash('راجع بيانات موقع التخزين.', 'danger')
    return redirect(url_for('storage_locations'))


@app.route('/storage-stocks', methods=['GET', 'POST'])
@login_required
@warehouse_required
def storage_stocks():
    form = ProductLocationStockForm()
    populate_location_stock_choices(form)
    if form.validate_on_submit():
        location = StorageLocation.query.get(form.location_id.data)
        if not can_access_branch(location.branch_id if location else None):
            flash('ليس لديك صلاحية تعديل كميات هذا الموقع.', 'danger')
            return redirect(url_for('storage_stocks'))
        stock = ProductLocationStock.query.filter_by(product_id=form.product_id.data, location_id=form.location_id.data).first()
        if not stock:
            stock = ProductLocationStock(product_id=form.product_id.data, location_id=form.location_id.data, quantity=0)
            db.session.add(stock)
            db.session.flush()
        old_location_qty = stock.quantity
        new_location_qty = form.quantity.data
        delta = new_location_qty - old_location_qty
        stock.quantity = new_location_qty
        stock.notes = form.notes.data

        inventory = Inventory.query.filter_by(product_id=form.product_id.data).first()
        if not inventory:
            inventory = Inventory(product_id=form.product_id.data, quantity=0)
            db.session.add(inventory)
            db.session.flush()
        before = inventory.quantity
        inventory.quantity += delta
        db.session.add(InventoryTransaction(
            product_id=form.product_id.data,
            quantity_before=before,
            quantity_change=delta,
            quantity_after=inventory.quantity,
            transaction_type='location_stock_adjustment',
            reference_type='storage_location',
            reference_id=form.location_id.data,
            notes=f'تعديل كمية موقع تخزين من {old_location_qty} إلى {new_location_qty}',
            user_id=current_user.id
        ))
        audit_log('update_location_stock', 'ProductLocationStock', stock.id, f'تعديل كمية موقع تخزين بفارق {delta}')
        commit_with_retry()
        flash('تم تحديث كمية الصنف داخل موقع التخزين وتحديث إجمالي المخزون.', 'success')
        return redirect(url_for('storage_stocks'))

    stocks_query = ProductLocationStock.query.options(
        joinedload(ProductLocationStock.product),
        joinedload(ProductLocationStock.location).joinedload(StorageLocation.branch)
    ).join(StorageLocation, ProductLocationStock.location_id == StorageLocation.id)
    if scoped_branch_id():
        stocks_query = stocks_query.filter(StorageLocation.branch_id == scoped_branch_id())
    stocks_page = stocks_query.order_by(ProductLocationStock.updated_at.desc()).paginate(page=request.args.get('page', 1, type=int), per_page=50, error_out=False)
    return render_template('storage_locations/stocks.html', form=form, stocks_page=stocks_page)


@app.route('/employees')
@login_required
@warehouse_required
def employees():
    employees_query = Employee.query.options(joinedload(Employee.branch))
    if scoped_branch_id():
        employees_query = employees_query.filter(Employee.branch_id == scoped_branch_id())
    employees = employees_query.order_by(Employee.is_active.desc(), Employee.name).all()
    form = EmployeeForm()
    populate_employee_form_choices(form)
    return render_template('employees/index.html', employees=employees, form=form)


@app.route('/employees/add', methods=['POST'])
@login_required
@warehouse_required
def add_employee():
    form = EmployeeForm()
    populate_employee_form_choices(form)
    if form.validate_on_submit():
        if not can_access_branch(form.branch_id.data):
            flash('ليس لديك صلاحية إضافة موظف لهذا الفرع.', 'danger')
            return redirect(url_for('employees'))
        employee_code = (form.employee_code.data or '').strip()
        if Employee.query.filter(func.lower(Employee.employee_code) == employee_code.lower()).first():
            flash('كود الموظف موجود بالفعل.', 'danger')
            return redirect(url_for('employees'))
        employee = Employee(
            employee_code=employee_code,
            name=form.name.data,
            branch_id=form.branch_id.data if form.branch_id.data else None,
            department=form.department.data,
            job_title=form.job_title.data,
            phone=form.phone.data,
            email=form.email.data,
            is_active=form.is_active.data,
            notes=form.notes.data
        )
        db.session.add(employee)
        commit_with_retry()
        flash('تم إضافة الموظف بنجاح.', 'success')
    else:
        flash('راجع بيانات الموظف المدخلة.', 'danger')
    return redirect(url_for('employees'))


@app.route('/employees/edit/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def edit_employee(id):
    employee = Employee.query.get_or_404(id)
    if not can_access_employee(employee):
        flash('ليس لديك صلاحية تعديل هذا الموظف.', 'danger')
        return redirect(url_for('employees'))
    form = EmployeeForm()
    populate_employee_form_choices(form)
    if form.validate_on_submit():
        if not can_access_branch(form.branch_id.data):
            flash('ليس لديك صلاحية نقل الموظف لهذا الفرع.', 'danger')
            return redirect(url_for('employees'))
        employee_code = (form.employee_code.data or '').strip()
        duplicate = Employee.query.filter(func.lower(Employee.employee_code) == employee_code.lower(), Employee.id != id).first()
        if duplicate:
            flash('كود الموظف موجود بالفعل.', 'danger')
            return redirect(url_for('employees'))
        employee.employee_code = employee_code
        employee.name = form.name.data
        employee.branch_id = form.branch_id.data if form.branch_id.data else None
        employee.department = form.department.data
        employee.job_title = form.job_title.data
        employee.phone = form.phone.data
        employee.email = form.email.data
        employee.is_active = form.is_active.data
        employee.notes = form.notes.data
        commit_with_retry()
        flash('تم تحديث بيانات الموظف.', 'success')
    else:
        flash('راجع بيانات الموظف المدخلة.', 'danger')
    return redirect(url_for('employees'))


@app.route('/inventory/issue', methods=['GET', 'POST'])
@login_required
@warehouse_required
def stock_issue():
    form = StockIssueForm()
    products = Product.query.order_by(Product.name).all()
    employees_query = Employee.query.options(joinedload(Employee.branch)).filter_by(is_active=True)
    if scoped_branch_id():
        employees_query = employees_query.filter(Employee.branch_id == scoped_branch_id())
    employees_list = employees_query.order_by(Employee.name).all()
    form.product_id.choices = [(p.id, p.name) for p in products]
    form.employee_id.choices = [
        (e.id, f'{e.employee_code or "-"} - {e.name} - {e.branch.name if e.branch else "بدون فرع"}')
        for e in employees_list
    ]

    if not products:
        flash('لا توجد منتجات للصرف.', 'warning')
    if not employees_list:
        flash('أضف موظفين أولاً قبل صرف الأصناف.', 'warning')

    recent_issues = StockIssue.query.order_by(StockIssue.created_at.desc()).limit(50).all()
    if form.validate_on_submit():
        inventory = Inventory.query.filter_by(product_id=form.product_id.data).first()
        employee = Employee.query.get(form.employee_id.data)
        if not can_access_employee(employee):
            flash('ليس لديك صلاحية صرف عهدة لهذا الموظف أو الفرع.', 'danger')
            return render_template('inventory/issue.html', form=form, recent_issues=recent_issues)
        if not inventory or inventory.quantity < form.quantity.data:
            flash('الكمية المطلوبة غير متاحة في المخزون.', 'danger')
            return render_template('inventory/issue.html', form=form, recent_issues=recent_issues)
        if not employee:
            flash('الموظف المحدد غير موجود.', 'danger')
            return render_template('inventory/issue.html', form=form, recent_issues=recent_issues)

        before = inventory.quantity
        inventory.quantity -= form.quantity.data
        issue = StockIssue(
            product_id=form.product_id.data,
            employee_id=form.employee_id.data,
            quantity=form.quantity.data,
            issue_date=form.issue_date.data,
            purpose=form.purpose.data,
            notes=form.notes.data,
            user_id=current_user.id
        )
        transaction = InventoryTransaction(
            product_id=form.product_id.data,
            quantity_before=before,
            quantity_change=-form.quantity.data,
            quantity_after=inventory.quantity,
            transaction_type='issue',
            reference_type='stock_issue',
            notes=f'صرف لموظف: {employee.name}',
            user_id=current_user.id
        )
        db.session.add(issue)
        db.session.add(transaction)
        commit_with_retry()
        transaction.reference_id = issue.id
        commit_with_retry()
        flash('تم صرف الصنف وتحديث المخزون.', 'success')
        return redirect(url_for('stock_issue'))

    return render_template('inventory/issue.html', form=form, recent_issues=recent_issues)


@app.route('/inventory/damage', methods=['GET', 'POST'])
@login_required
@warehouse_required
def damage_record():
    form = DamageRecordForm()
    products = Product.query.order_by(Product.name).all()
    form.product_id.choices = [(p.id, p.name) for p in products]
    recent_damage = DamageRecord.query.order_by(DamageRecord.created_at.desc()).limit(50).all()

    if form.validate_on_submit():
        inventory = Inventory.query.filter_by(product_id=form.product_id.data).first()
        if not inventory or inventory.quantity < form.quantity.data:
            flash('كمية الهالك أكبر من المتاح في المخزون.', 'danger')
            return render_template('inventory/damage.html', form=form, recent_damage=recent_damage)

        before = inventory.quantity
        inventory.quantity -= form.quantity.data
        damage = DamageRecord(
            product_id=form.product_id.data,
            quantity=form.quantity.data,
            damage_date=form.damage_date.data,
            reason=form.reason.data,
            responsibility=form.responsibility.data,
            notes=form.notes.data,
            user_id=current_user.id
        )
        transaction = InventoryTransaction(
            product_id=form.product_id.data,
            quantity_before=before,
            quantity_change=-form.quantity.data,
            quantity_after=inventory.quantity,
            transaction_type='damage',
            reference_type='damage_record',
            notes=form.reason.data or 'تسجيل هالك',
            user_id=current_user.id
        )
        db.session.add(damage)
        db.session.add(transaction)
        commit_with_retry()
        transaction.reference_id = damage.id
        commit_with_retry()
        flash('تم تسجيل الهالك وتحديث المخزون.', 'success')
        return redirect(url_for('damage_record'))

    return render_template('inventory/damage.html', form=form, recent_damage=recent_damage)


@app.route('/inventory/issue-requests', methods=['GET', 'POST'])
@login_required
def issue_requests():
    form = StockIssueRequestForm()
    populate_product_employee_choices(form)
    if form.validate_on_submit():
        if not has_permission('issue_requests.create'):
            flash('ليس لديك صلاحية إنشاء طلب صرف.', 'danger')
            return redirect(url_for('issue_requests'))
        employee = Employee.query.get(form.employee_id.data)
        if not can_access_employee(employee):
            flash('ليس لديك صلاحية إنشاء طلب صرف لهذا الموظف أو الفرع.', 'danger')
            return redirect(url_for('issue_requests'))
        request_number = generate_sequence(StockIssueRequest, 'request_number', 'REQ')
        issue_request = StockIssueRequest(
            request_number=request_number,
            product_id=form.product_id.data,
            employee_id=form.employee_id.data,
            quantity=form.quantity.data,
            purpose=form.purpose.data,
            notes=form.notes.data,
            requested_by_id=current_user.id
        )
        db.session.add(issue_request)
        audit_log('create_issue_request', 'StockIssueRequest', None, f'طلب صرف {request_number}')
        commit_with_retry()
        flash('تم إنشاء طلب الصرف وفي انتظار الاعتماد.', 'success')
        return redirect(url_for('issue_requests'))

    status = request.args.get('status', 'all')
    query = StockIssueRequest.query.options(
        joinedload(StockIssueRequest.product),
        joinedload(StockIssueRequest.employee).joinedload(Employee.branch),
        joinedload(StockIssueRequest.requested_by),
        joinedload(StockIssueRequest.approved_by),
        joinedload(StockIssueRequest.executed_by)
    )
    if status != 'all':
        query = query.filter_by(status=status)
    if scoped_branch_id():
        query = query.join(Employee, StockIssueRequest.employee_id == Employee.id).filter(Employee.branch_id == scoped_branch_id())
    elif not (has_permission('issue_requests.approve') or has_permission('issue_requests.execute') or has_permission('audit.view')):
        query = query.filter_by(requested_by_id=current_user.id)
    requests_page = query.order_by(StockIssueRequest.requested_at.desc()).paginate(
        page=request.args.get('page', 1, type=int),
        per_page=25,
        error_out=False
    )
    return render_template('inventory/issue_requests.html', form=form, requests_page=requests_page, status=status, can_approve=can_approve_requests(), can_execute=can_manage_warehouse())


@app.route('/inventory/issue-requests/<int:id>/approve', methods=['POST'])
@login_required
@roles_required('admin', 'approver', 'warehouse_manager')
def approve_issue_request(id):
    issue_request = StockIssueRequest.query.get_or_404(id)
    if not can_access_employee(issue_request.employee):
        flash('ليس لديك صلاحية اعتماد طلب خارج نطاق فرعك.', 'danger')
        return redirect(url_for('issue_requests'))
    if issue_request.status != 'pending':
        flash('لا يمكن اعتماد طلب ليس في حالة انتظار.', 'warning')
        return redirect(url_for('issue_requests'))
    signature = require_e_signature('اعتماد طلب صرف')
    if not signature:
        return redirect(url_for('issue_requests'))
    issue_request.status = 'approved'
    issue_request.approved_by_id = current_user.id
    issue_request.approved_at = datetime.utcnow()
    issue_request.approved_signature = signature
    audit_log('approve_issue_request', 'StockIssueRequest', id, issue_request.request_number)
    commit_with_retry()
    flash('تم اعتماد طلب الصرف.', 'success')
    return redirect(url_for('issue_requests', status='approved'))


@app.route('/inventory/issue-requests/<int:id>/reject', methods=['POST'])
@login_required
@roles_required('admin', 'approver', 'warehouse_manager')
def reject_issue_request(id):
    issue_request = StockIssueRequest.query.get_or_404(id)
    if not can_access_employee(issue_request.employee):
        flash('ليس لديك صلاحية رفض طلب خارج نطاق فرعك.', 'danger')
        return redirect(url_for('issue_requests'))
    if issue_request.status not in ('pending', 'approved'):
        flash('لا يمكن رفض هذا الطلب.', 'warning')
        return redirect(url_for('issue_requests'))
    signature = require_e_signature('رفض طلب صرف')
    if not signature:
        return redirect(url_for('issue_requests'))
    issue_request.status = 'rejected'
    issue_request.rejected_by_id = current_user.id
    issue_request.rejected_at = datetime.utcnow()
    issue_request.rejected_signature = signature
    issue_request.rejection_reason = request.form.get('rejection_reason')
    audit_log('reject_issue_request', 'StockIssueRequest', id, issue_request.request_number)
    commit_with_retry()
    flash('تم رفض طلب الصرف.', 'success')
    return redirect(url_for('issue_requests'))


@app.route('/inventory/issue-requests/<int:id>/execute', methods=['POST'])
@login_required
@roles_required('admin', 'warehouse_manager')
def execute_issue_request(id):
    issue_request = StockIssueRequest.query.options(joinedload(StockIssueRequest.employee)).get_or_404(id)
    if not can_access_employee(issue_request.employee):
        flash('ليس لديك صلاحية تنفيذ طلب خارج نطاق فرعك.', 'danger')
        return redirect(url_for('issue_requests'))
    if issue_request.status != 'approved':
        flash('لا يمكن تنفيذ طلب غير معتمد.', 'warning')
        return redirect(url_for('issue_requests'))
    signature = require_e_signature('تنفيذ طلب صرف')
    if not signature:
        return redirect(url_for('issue_requests', status='approved'))
    inventory = Inventory.query.filter_by(product_id=issue_request.product_id).first()
    if not inventory or inventory.quantity < issue_request.quantity:
        flash('الكمية غير متاحة في المخزون.', 'danger')
        return redirect(url_for('issue_requests', status='approved'))
    before = inventory.quantity
    inventory.quantity -= issue_request.quantity
    issue = StockIssue(
        product_id=issue_request.product_id,
        employee_id=issue_request.employee_id,
        quantity=issue_request.quantity,
        issue_date=datetime.now().date(),
        purpose=issue_request.purpose,
        notes=issue_request.notes,
        user_id=current_user.id
    )
    db.session.add(issue)
    db.session.flush()
    transaction = InventoryTransaction(
        product_id=issue_request.product_id,
        quantity_before=before,
        quantity_change=-issue_request.quantity,
        quantity_after=inventory.quantity,
        transaction_type='issue_request_execute',
        reference_type='stock_issue_request',
        reference_id=issue_request.id,
        notes=f'تنفيذ طلب صرف {issue_request.request_number}',
        user_id=current_user.id
    )
    db.session.add(transaction)
    issue_request.status = 'executed'
    issue_request.executed_by_id = current_user.id
    issue_request.executed_at = datetime.utcnow()
    issue_request.executed_signature = signature
    issue_request.stock_issue_id = issue.id
    audit_log('execute_issue_request', 'StockIssueRequest', id, issue_request.request_number)
    commit_with_retry()
    flash('تم تنفيذ طلب الصرف وتحديث المخزون.', 'success')
    return redirect(url_for('issue_requests', status='executed'))


@app.route('/inventory/returns', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'warehouse_manager')
def employee_returns():
    form = EmployeeReturnForm()
    populate_product_employee_choices(form)
    if form.validate_on_submit():
        employee = Employee.query.get(form.employee_id.data)
        if not can_access_employee(employee):
            flash('ليس لديك صلاحية تسجيل مرتجع لهذا الموظف أو الفرع.', 'danger')
            return redirect(url_for('employee_returns'))
        signature = require_e_signature('تسجيل مرتجع عهدة')
        if not signature:
            return redirect(url_for('employee_returns'))
        inventory = Inventory.query.filter_by(product_id=form.product_id.data).first()
        if not inventory:
            inventory = Inventory(product_id=form.product_id.data, quantity=0)
            db.session.add(inventory)
            db.session.flush()
        before = inventory.quantity
        if form.condition.data == 'usable':
            inventory.quantity += form.quantity.data
            after = inventory.quantity
            transaction_type = 'employee_return'
        else:
            after = inventory.quantity
            transaction_type = 'employee_return_review'
        return_record = EmployeeReturn(
            return_number=generate_sequence(EmployeeReturn, 'return_number', 'RET'),
            product_id=form.product_id.data,
            employee_id=form.employee_id.data,
            quantity=form.quantity.data,
            return_date=form.return_date.data,
            condition=form.condition.data,
            notes=form.notes.data,
            user_id=current_user.id,
            signed_signature=signature
        )
        db.session.add(return_record)
        db.session.add(InventoryTransaction(
            product_id=form.product_id.data,
            quantity_before=before,
            quantity_change=form.quantity.data if form.condition.data == 'usable' else 0,
            quantity_after=after,
            transaction_type=transaction_type,
            reference_type='employee_return',
            notes=f'مرتجع عهدة {return_record.return_number}',
            user_id=current_user.id
        ))
        audit_log('create_employee_return', 'EmployeeReturn', None, return_record.return_number)
        commit_with_retry()
        flash('تم تسجيل مرتجع العهدة.', 'success')
        return redirect(url_for('employee_returns'))
    returns_query = EmployeeReturn.query.options(joinedload(EmployeeReturn.product), joinedload(EmployeeReturn.employee).joinedload(Employee.branch), joinedload(EmployeeReturn.user)).join(Employee, EmployeeReturn.employee_id == Employee.id)
    if scoped_branch_id():
        returns_query = returns_query.filter(Employee.branch_id == scoped_branch_id())
    returns_page = returns_query.order_by(EmployeeReturn.created_at.desc()).paginate(page=request.args.get('page', 1, type=int), per_page=25, error_out=False)
    return render_template('inventory/returns.html', form=form, returns_page=returns_page)


@app.route('/inventory/stocktakes', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def stocktakes():
    form = StocktakeForm()
    branches = Branch.query.filter_by(is_active=True).order_by(Branch.name).all()
    form.branch_id.choices = [(0, 'بدون فرع / كل الفروع')] + [(branch.id, f'{branch.name} ({branch.code})') for branch in branches]
    if form.validate_on_submit():
        if scoped_branch_id() and form.branch_id.data != scoped_branch_id():
            flash('ليس لديك صلاحية إنشاء جرد خارج نطاق فرعك.', 'danger')
            return redirect(url_for('stocktakes'))
        stocktake = Stocktake(
            count_number=generate_sequence(Stocktake, 'count_number', 'CNT'),
            branch_id=form.branch_id.data if form.branch_id.data else None,
            count_date=form.count_date.data,
            notes=form.notes.data,
            created_by_id=current_user.id
        )
        db.session.add(stocktake)
        audit_log('create_stocktake', 'Stocktake', None, stocktake.count_number)
        commit_with_retry()
        flash('تم إنشاء جرد جديد. أضف الأصناف والكمية الفعلية.', 'success')
        return redirect(url_for('view_stocktake', id=stocktake.id))
    stocktakes_query = Stocktake.query.options(joinedload(Stocktake.branch), joinedload(Stocktake.created_by))
    if scoped_branch_id():
        stocktakes_query = stocktakes_query.filter(Stocktake.branch_id == scoped_branch_id())
    stocktakes_page = stocktakes_query.order_by(Stocktake.created_at.desc()).paginate(page=request.args.get('page', 1, type=int), per_page=25, error_out=False)
    return render_template('inventory/stocktakes.html', form=form, stocktakes_page=stocktakes_page)


@app.route('/inventory/stocktakes/<int:id>', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def view_stocktake(id):
    stocktake = Stocktake.query.options(joinedload(Stocktake.items).joinedload(StocktakeItem.product)).get_or_404(id)
    if not can_access_branch(stocktake.branch_id):
        flash('ليس لديك صلاحية فتح جرد خارج نطاق فرعك.', 'danger')
        return redirect(url_for('stocktakes'))
    form = StocktakeItemForm()
    products = Product.query.order_by(Product.name).all()
    form.product_id.choices = [(p.id, p.name) for p in products]
    if form.validate_on_submit() and stocktake.status == 'draft':
        inventory = Inventory.query.filter_by(product_id=form.product_id.data).first()
        system_quantity = inventory.quantity if inventory else 0
        counted_quantity = form.counted_quantity.data
        existing = StocktakeItem.query.filter_by(stocktake_id=stocktake.id, product_id=form.product_id.data).first()
        if existing:
            existing.system_quantity = system_quantity
            existing.counted_quantity = counted_quantity
            existing.variance = counted_quantity - system_quantity
            existing.notes = form.notes.data
        else:
            db.session.add(StocktakeItem(
                stocktake_id=stocktake.id,
                product_id=form.product_id.data,
                system_quantity=system_quantity,
                counted_quantity=counted_quantity,
                variance=counted_quantity - system_quantity,
                notes=form.notes.data
            ))
        audit_log('update_stocktake_item', 'Stocktake', stocktake.id, stocktake.count_number)
        commit_with_retry()
        flash('تم تسجيل بند الجرد.', 'success')
        return redirect(url_for('view_stocktake', id=stocktake.id))
    return render_template('inventory/stocktake_view.html', stocktake=stocktake, form=form)


@app.route('/inventory/stocktakes/<int:id>/submit', methods=['POST'])
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def submit_stocktake(id):
    stocktake = Stocktake.query.get_or_404(id)
    if not can_access_branch(stocktake.branch_id):
        flash('ليس لديك صلاحية إرسال هذا الجرد للاعتماد.', 'danger')
        return redirect(url_for('stocktakes'))
    if stocktake.status != 'draft':
        flash('لا يمكن إرسال هذا الجرد للاعتماد.', 'warning')
        return redirect(url_for('view_stocktake', id=id))
    if not stocktake.items:
        flash('أضف بنود الجرد قبل الإرسال للاعتماد.', 'warning')
        return redirect(url_for('view_stocktake', id=id))
    stocktake.status = 'submitted'
    audit_log('submit_stocktake', 'Stocktake', stocktake.id, stocktake.count_number)
    commit_with_retry()
    flash('تم إرسال الجرد للاعتماد.', 'success')
    return redirect(url_for('view_stocktake', id=id))


@app.route('/inventory/stocktakes/<int:id>/reject', methods=['POST'])
@login_required
@roles_required('admin', 'warehouse_manager')
def reject_stocktake(id):
    stocktake = Stocktake.query.get_or_404(id)
    if not can_access_branch(stocktake.branch_id):
        flash('ليس لديك صلاحية رفض جرد خارج نطاق فرعك.', 'danger')
        return redirect(url_for('stocktakes'))
    if stocktake.status != 'submitted':
        flash('لا يمكن رفض جرد غير مرسل للاعتماد.', 'warning')
        return redirect(url_for('view_stocktake', id=id))
    signature = require_e_signature('رفض جرد')
    if not signature:
        return redirect(url_for('view_stocktake', id=id))
    stocktake.status = 'rejected'
    stocktake.rejected_by_id = current_user.id
    stocktake.rejected_at = datetime.utcnow()
    stocktake.rejected_signature = signature
    stocktake.rejection_reason = request.form.get('rejection_reason')
    audit_log('reject_stocktake', 'Stocktake', stocktake.id, stocktake.count_number)
    commit_with_retry()
    flash('تم رفض الجرد. يمكن نسخه أو إنشاء جرد جديد بعد المراجعة.', 'success')
    return redirect(url_for('view_stocktake', id=id))


@app.route('/inventory/stocktakes/<int:id>/approve', methods=['POST'])
@login_required
@roles_required('admin', 'warehouse_manager')
def approve_stocktake(id):
    stocktake = Stocktake.query.options(joinedload(Stocktake.items)).get_or_404(id)
    if not can_access_branch(stocktake.branch_id):
        flash('ليس لديك صلاحية اعتماد جرد خارج نطاق فرعك.', 'danger')
        return redirect(url_for('stocktakes'))
    if stocktake.status != 'submitted':
        flash('لا يمكن اعتماد الجرد إلا بعد إرساله للاعتماد.', 'warning')
        return redirect(url_for('view_stocktake', id=id))
    signature = require_e_signature('اعتماد وتسوية جرد')
    if not signature:
        return redirect(url_for('view_stocktake', id=id))
    for item in stocktake.items:
        if item.variance == 0:
            continue
        inventory = Inventory.query.filter_by(product_id=item.product_id).first()
        if not inventory:
            inventory = Inventory(product_id=item.product_id, quantity=0)
            db.session.add(inventory)
            db.session.flush()
        before = inventory.quantity
        inventory.quantity = item.counted_quantity
        db.session.add(InventoryTransaction(
            product_id=item.product_id,
            quantity_before=before,
            quantity_change=item.variance,
            quantity_after=inventory.quantity,
            transaction_type='stocktake_adjustment',
            reference_type='stocktake',
            reference_id=stocktake.id,
            notes=f'تسوية جرد {stocktake.count_number}',
            user_id=current_user.id
        ))
    stocktake.status = 'approved'
    stocktake.approved_by_id = current_user.id
    stocktake.approved_at = datetime.utcnow()
    stocktake.approved_signature = signature
    audit_log('approve_stocktake', 'Stocktake', stocktake.id, stocktake.count_number)
    commit_with_retry()
    flash('تم اعتماد الجرد وتسوية فروقات المخزون.', 'success')
    return redirect(url_for('view_stocktake', id=id))

# Reports routes
@app.route('/reports')
@login_required
@report_required
def reports():
    categories = Category.query.order_by(Category.name).all()
    branches_query = Branch.query.filter_by(is_active=True)
    if scoped_branch_id():
        branches_query = branches_query.filter(Branch.id == scoped_branch_id())
    branches = branches_query.order_by(Branch.name).all()
    return render_template('reports/index.html', categories=categories, branches=branches)

@app.route('/reports/sales')
@login_required
@admin_required
def sales_report():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    sales = Sale.query.filter(Sale.sale_date.between(start_date, end_date)).order_by(Sale.sale_date.desc()).all()

    total_sales = sum(sale.total_amount for sale in sales)
    total_profit = sum(sale.profit for sale in sales if sale.profit)
    average_sale = total_sales / len(sales) if sales else 0

    sales_by_date = OrderedDict()
    sales_by_payment = defaultdict(lambda: {'count': 0, 'total': 0})
    sales_by_customer = defaultdict(lambda: {'count': 0, 'total': 0})
    for sale in sales:
        date_key = sale.sale_date.strftime('%Y-%m-%d')
        day_bucket = sales_by_date.setdefault(date_key, {'count': 0, 'total': 0})
        day_bucket['count'] += 1
        day_bucket['total'] += sale.total_amount
        payment_bucket = sales_by_payment[sale.payment_method]
        payment_bucket['count'] += 1
        payment_bucket['total'] += sale.total_amount
        customer_key = sale.customer.name if sale.customer else 'غير محدد'
        customer_bucket = sales_by_customer[customer_key]
        customer_bucket['count'] += 1
        customer_bucket['total'] += sale.total_amount

    completed_sales = [sale for sale in sales if sale.status == 'completed']
    cancelled_sales = [sale for sale in sales if sale.status == 'cancelled']
    majel_sales = [sale for sale in sales if sale.payment_method == 'majel' and sale.status != 'cancelled']
    completed_total = sum(sale.total_amount for sale in completed_sales)
    cancelled_total = sum(sale.total_amount for sale in cancelled_sales)
    majel_total = sum(sale.total_amount for sale in majel_sales)
    majel_due_total = sum(sale.due_amount for sale in majel_sales)
    net_after_due = total_sales - majel_due_total
    
    report_type = request.args.get('report_type', 'summary')
    return render_template('reports/sales_report.html', 
                          sales=sales, 
                          total_sales=total_sales,
                          total_profit=total_profit,
                          average_sale=average_sale,
                          start_date=start_date,
                          end_date=end_date,
                          now=datetime.now(),
                          completed_sales=completed_sales,
                          cancelled_sales=cancelled_sales,
                          majel_sales=majel_sales,
                          completed_total=completed_total,
                          cancelled_total=cancelled_total,
                          majel_total=majel_total,
                          majel_due_total=majel_due_total,
                          net_after_due=net_after_due,
                          sales_by_date=sales_by_date,
                          sales_by_payment=sales_by_payment,
                          sales_by_customer=sales_by_customer,
                          report_type=report_type)

@app.route('/reports/purchases')
@login_required
@report_required
def purchases_report():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    purchases = Purchase.query.filter(Purchase.purchase_date.between(start_date, end_date)).order_by(Purchase.purchase_date.desc()).all()
    
    total_purchases = sum(purchase.total_amount for purchase in purchases)
    average_purchase = total_purchases / len(purchases) if purchases else 0

    purchases_by_date = OrderedDict()
    purchases_by_payment = defaultdict(lambda: {'count': 0, 'total': 0})
    purchases_by_supplier = defaultdict(lambda: {'count': 0, 'total': 0})
    for purchase in purchases:
        date_key = purchase.purchase_date.strftime('%Y-%m-%d')
        day_bucket = purchases_by_date.setdefault(date_key, {'count': 0, 'total': 0})
        day_bucket['count'] += 1
        day_bucket['total'] += purchase.total_amount
        payment_bucket = purchases_by_payment[purchase.payment_method]
        payment_bucket['count'] += 1
        payment_bucket['total'] += purchase.total_amount
        supplier_key = purchase.supplier.name if purchase.supplier else 'غير محدد'
        supplier_bucket = purchases_by_supplier[supplier_key]
        supplier_bucket['count'] += 1
        supplier_bucket['total'] += purchase.total_amount

    completed_purchases = [p for p in purchases if p.status == 'completed']
    cancelled_purchases = [p for p in purchases if p.status == 'cancelled']
    majel_purchases = [p for p in purchases if p.payment_method == 'majel' and p.status != 'cancelled']
    completed_total = sum(p.total_amount for p in completed_purchases)
    cancelled_total = sum(p.total_amount for p in cancelled_purchases)
    majel_total = sum(p.total_amount for p in majel_purchases)
    majel_due_total = sum((p.total_amount - p.amount_paid) for p in majel_purchases)
    
    report_type = request.args.get('report_type', 'summary')
    return render_template('reports/purchases_report.html', 
                          purchases=purchases, 
                          total_purchases=total_purchases,
                          average_purchase=average_purchase,
                          start_date=start_date,
                          end_date=end_date,
                          now=datetime.now(),
                          purchases_by_date=purchases_by_date,
                          purchases_by_payment=purchases_by_payment,
                          purchases_by_supplier=purchases_by_supplier,
                          completed_purchases=completed_purchases,
                          cancelled_purchases=cancelled_purchases,
                          majel_purchases=majel_purchases,
                          completed_total=completed_total,
                          cancelled_total=cancelled_total,
                          majel_total=majel_total,
                          majel_due_total=majel_due_total,
                          report_type=report_type)

@app.route('/reports/inventory')
@login_required
@report_required
def inventory_report():
    status = request.args.get('status', 'all')
    category_id = request.args.get('category_id', type=int)
    query = db.session.query(Product, Inventory).join(
        Inventory, Product.id == Inventory.product_id
    )
    if status == 'low':
        query = query.filter(Inventory.quantity <= Product.min_quantity, Inventory.quantity > 0)
    elif status == 'out':
        query = query.filter(Inventory.quantity <= 0)
    if category_id:
        query = query.filter(Product.category_id == category_id)
    inventory_items = query.order_by(Product.name).all()
    
    total_value = sum(item.quantity * item.product.purchase_price for _, item in inventory_items)
    categories = Category.query.order_by(Category.name).all()
    
    return render_template('reports/inventory_report.html', 
                          inventory=inventory_items, 
                          total_value=total_value,
                          now=datetime.now(),
                          status=status,
                          category_id=category_id,
                          categories=categories)


@app.route('/reports/issues')
@login_required
@report_required
def issues_report():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    branch_id = request.args.get('branch_id', type=int)
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
    query = StockIssue.query.options(
        joinedload(StockIssue.product),
        joinedload(StockIssue.employee).joinedload(Employee.branch),
        joinedload(StockIssue.user)
    ).join(Employee, StockIssue.employee_id == Employee.id).filter(StockIssue.issue_date.between(start_date, end_date))
    if scoped_branch_id():
        branch_id = scoped_branch_id()
        query = query.filter(Employee.branch_id == branch_id)
    elif branch_id:
        query = query.filter(Employee.branch_id == branch_id)
    issues = query.order_by(StockIssue.issue_date.desc(), StockIssue.id.desc()).all()
    total_quantity = sum(issue.quantity for issue in issues)
    total_value = sum(issue.quantity * (issue.product.purchase_price or 0) for issue in issues if issue.product)
    by_employee = defaultdict(int)
    by_product = defaultdict(int)
    by_branch = defaultdict(int)
    for issue in issues:
        employee_label = f'{issue.employee.employee_code or "-"} - {issue.employee.name}' if issue.employee else 'غير محدد'
        product_label = issue.product.name if issue.product else 'غير محدد'
        branch_label = issue.employee.branch.name if issue.employee and issue.employee.branch else 'بدون فرع'
        by_employee[employee_label] += issue.quantity
        by_product[product_label] += issue.quantity
        by_branch[branch_label] += issue.quantity
    branches_query = Branch.query.filter_by(is_active=True)
    if scoped_branch_id():
        branches_query = branches_query.filter(Branch.id == scoped_branch_id())
    branches = branches_query.order_by(Branch.name).all()
    return render_template(
        'reports/issues_report.html',
        issues=issues,
        total_quantity=total_quantity,
        total_value=total_value,
        by_employee=sorted(by_employee.items(), key=lambda item: item[1], reverse=True),
        by_product=sorted(by_product.items(), key=lambda item: item[1], reverse=True),
        by_branch=sorted(by_branch.items(), key=lambda item: item[1], reverse=True),
        branches=branches,
        branch_id=branch_id,
        start_date=start_date,
        end_date=end_date,
        now=datetime.now()
    )


@app.route('/reports/damage')
@login_required
@report_required
def damage_report():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if isinstance(start_date, str) else start_date
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date() if isinstance(end_date, str) else end_date
    records = DamageRecord.query.filter(DamageRecord.damage_date.between(start_date, end_date)).order_by(DamageRecord.damage_date.desc()).all()
    total_quantity = sum(record.quantity for record in records)
    total_value = sum(record.quantity * (record.product.purchase_price or 0) for record in records)
    by_product = defaultdict(int)
    by_reason = defaultdict(int)
    for record in records:
        by_product[record.product.name] += record.quantity
        by_reason[record.reason or 'غير محدد'] += record.quantity
    return render_template(
        'reports/damage_report.html',
        records=records,
        total_quantity=total_quantity,
        total_value=total_value,
        by_product=sorted(by_product.items(), key=lambda item: item[1], reverse=True),
        by_reason=sorted(by_reason.items(), key=lambda item: item[1], reverse=True),
        start_date=start_date,
        end_date=end_date,
        now=datetime.now()
    )


@app.route('/reports/inventory/export.xlsx')
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def export_inventory_excel():
    rows = []
    items = db.session.query(Product, Inventory).join(Inventory, Product.id == Inventory.product_id).order_by(Product.name).all()
    for product, item in items:
        rows.append([
            product.name,
            product.sku or '',
            product.barcode or '',
            product.category.name if product.category else '',
            item.quantity,
            product.min_quantity,
            product.purchase_price or 0,
            item.quantity * (product.purchase_price or 0)
        ])
    return excel_response(
        'inventory_report.xlsx',
        ['الصنف', 'SKU', 'الباركود', 'التصنيف', 'الكمية', 'الحد الأدنى', 'سعر الشراء', 'القيمة'],
        rows,
        'تقرير المخزون'
    )


@app.route('/reports/inventory/export.pdf')
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def export_inventory_pdf():
    items = db.session.query(Product, Inventory).join(Inventory, Product.id == Inventory.product_id).order_by(Product.name).all()
    rows = []
    total_qty = 0
    total_value = 0
    for product, item in items:
        value = item.quantity * (product.purchase_price or 0)
        total_qty += item.quantity
        total_value += value
        rows.append([product.name, product.sku or '', product.barcode or '', product.category.name if product.category else '', item.quantity, product.min_quantity, f'{product.purchase_price or 0:.2f}', f'{value:.2f}'])
    return pdf_response(
        'inventory_report.pdf',
        'تقرير المخزون',
        ['الصنف', 'SKU', 'الباركود', 'التصنيف', 'الكمية', 'الحد الأدنى', 'سعر الشراء', 'القيمة'],
        rows,
        summary=[('عدد الأصناف', len(rows)), ('إجمالي الكمية', total_qty), ('قيمة المخزون', f'{total_value:.2f}')]
    )


@app.route('/reports/issues/export.xlsx')
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def export_issues_excel():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    issues_query = StockIssue.query.options(joinedload(StockIssue.product), joinedload(StockIssue.employee).joinedload(Employee.branch), joinedload(StockIssue.user)).join(Employee, StockIssue.employee_id == Employee.id).filter(StockIssue.issue_date.between(start_date, end_date))
    if scoped_branch_id():
        issues_query = issues_query.filter(Employee.branch_id == scoped_branch_id())
    issues = issues_query.order_by(StockIssue.issue_date.desc()).all()
    rows = []
    for issue in issues:
        rows.append([
            issue.issue_date.strftime('%Y-%m-%d'),
            issue.product.name if issue.product else '',
            issue.employee.employee_code if issue.employee else '',
            issue.employee.name if issue.employee else '',
            issue.employee.branch.name if issue.employee and issue.employee.branch else '',
            issue.quantity,
            issue.product.purchase_price if issue.product else 0,
            issue.quantity * (issue.product.purchase_price or 0) if issue.product else 0,
            issue.purpose or '',
            issue.user.username if issue.user else ''
        ])
    return excel_response(
        'issues_report.xlsx',
        ['التاريخ', 'الصنف', 'كود الموظف', 'الموظف', 'الفرع', 'الكمية', 'سعر الوحدة', 'القيمة', 'الغرض', 'المسجل'],
        rows,
        'تقرير الصرف الداخلي'
    )


@app.route('/reports/issues/export.pdf')
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def export_issues_pdf():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    issues_query = StockIssue.query.options(joinedload(StockIssue.product), joinedload(StockIssue.employee).joinedload(Employee.branch), joinedload(StockIssue.user)).join(Employee, StockIssue.employee_id == Employee.id).filter(StockIssue.issue_date.between(start_date_obj, end_date_obj))
    if scoped_branch_id():
        issues_query = issues_query.filter(Employee.branch_id == scoped_branch_id())
    issues = issues_query.order_by(StockIssue.issue_date.desc()).all()
    rows = []
    total_qty = 0
    total_value = 0
    for issue in issues:
        value = issue.quantity * (issue.product.purchase_price or 0) if issue.product else 0
        total_qty += issue.quantity
        total_value += value
        rows.append([issue.issue_date.strftime('%Y-%m-%d'), issue.product.name if issue.product else '', issue.employee.employee_code if issue.employee else '', issue.employee.name if issue.employee else '', issue.employee.branch.name if issue.employee and issue.employee.branch else '', issue.quantity, f'{value:.2f}', issue.purpose or ''])
    return pdf_response(
        'issues_report.pdf',
        'تقرير الصرف الداخلي',
        ['التاريخ', 'الصنف', 'كود الموظف', 'الموظف', 'الفرع', 'الكمية', 'القيمة', 'الغرض'],
        rows,
        subtitle=f'الفترة من {start_date} إلى {end_date}',
        summary=[('عدد العمليات', len(rows)), ('إجمالي الكمية', total_qty), ('القيمة التقديرية', f'{total_value:.2f}')]
    )


@app.route('/reports/damage/export.xlsx')
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def export_damage_excel():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    records = DamageRecord.query.options(joinedload(DamageRecord.product), joinedload(DamageRecord.user)).filter(DamageRecord.damage_date.between(start_date, end_date)).order_by(DamageRecord.damage_date.desc()).all()
    rows = []
    for record in records:
        rows.append([
            record.damage_date.strftime('%Y-%m-%d'),
            record.product.name if record.product else '',
            record.quantity,
            record.product.purchase_price if record.product else 0,
            record.quantity * (record.product.purchase_price or 0) if record.product else 0,
            record.reason or '',
            record.responsibility or '',
            record.user.username if record.user else '',
            record.notes or ''
        ])
    return excel_response(
        'damage_report.xlsx',
        ['التاريخ', 'الصنف', 'الكمية', 'سعر الوحدة', 'القيمة', 'السبب', 'المسؤولية', 'المسجل', 'ملاحظات'],
        rows,
        'تقرير الهالك'
    )

@app.route('/reports/damage/export.pdf')
@login_required
@roles_required('admin', 'warehouse_manager', 'auditor')
def export_damage_pdf():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    records = DamageRecord.query.options(joinedload(DamageRecord.product), joinedload(DamageRecord.user)).filter(DamageRecord.damage_date.between(start_date_obj, end_date_obj)).order_by(DamageRecord.damage_date.desc()).all()
    rows = []
    total_qty = 0
    total_value = 0
    for record in records:
        value = record.quantity * (record.product.purchase_price or 0) if record.product else 0
        total_qty += record.quantity
        total_value += value
        rows.append([record.damage_date.strftime('%Y-%m-%d'), record.product.name if record.product else '', record.quantity, f'{value:.2f}', record.reason or '', record.responsibility or '', record.user.username if record.user else ''])
    return pdf_response(
        'damage_report.pdf',
        'تقرير الهالك',
        ['التاريخ', 'الصنف', 'الكمية', 'القيمة', 'السبب', 'المسؤولية', 'المسجل'],
        rows,
        subtitle=f'الفترة من {start_date} إلى {end_date}',
        summary=[('عدد السجلات', len(rows)), ('إجمالي الكمية', total_qty), ('قيمة الهالك', f'{total_value:.2f}')]
    )


@app.route('/reports/top-selling', methods=['GET'])
@login_required
@admin_required
def top_selling_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    limit = request.args.get('limit', 10, type=int)

    # تحويل التواريخ
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # جلب المنتجات الأكثر مبيعاً
    query = db.session.query(
        Product,
        db.func.sum(SaleItem.quantity).label('total_quantity')
    ).join(SaleItem, SaleItem.product_id == Product.id
    ).join(Sale, Sale.id == SaleItem.sale_id)

    if start_date:
        query = query.filter(Sale.sale_date >= start_date)
    if end_date:
        query = query.filter(Sale.sale_date <= end_date)

    query = query.group_by(Product.id).order_by(db.desc('total_quantity')).limit(limit)
    results = query.all()

    return render_template('reports/top_selling.html', results=results, start_date=start_date, end_date=end_date)

@app.route('/reports/profit', methods=['GET'])
@login_required
@admin_required
def profit_report():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    group_by = request.args.get('group_by', 'day')

    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    items = db.session.query(SaleItem, Sale, Product).join(
        Sale, SaleItem.sale_id == Sale.id
    ).outerjoin(
        Product, SaleItem.product_id == Product.id
    ).filter(
        Sale.sale_date.between(start_date, end_date)
    ).all()

    total_profit = 0
    total_sales = 0
    grouped = OrderedDict()
    for sale_item, sale, product in items:
        sale_profit = 0
        if product and product.purchase_price is not None:
            sale_profit = (sale_item.price - product.purchase_price) * sale_item.quantity
        total_profit += sale_profit
        total_sales += sale_item.price * sale_item.quantity
        if group_by == 'week':
            week_label = f"{sale.sale_date.isocalendar()[0]}-{sale.sale_date.isocalendar()[1]}"
            key = f"أسبوع {week_label}"
        elif group_by == 'month':
            key = sale.sale_date.strftime('%Y-%m')
        else:
            key = sale.sale_date.strftime('%Y-%m-%d')
        bucket = grouped.setdefault(key, {'profit': 0, 'sales': 0})
        bucket['profit'] += sale_profit
        bucket['sales'] += sale_item.price * sale_item.quantity

    summary = [{
        'period': period,
        'sales': data['sales'],
        'profit': data['profit']
    } for period, data in grouped.items()]

    return render_template('reports/profit_report.html',
                          total_profit=total_profit,
                          total_sales=total_sales,
                          summary=summary,
                          start_date=start_date,
                          end_date=end_date,
                          group_by=group_by)


@app.route('/reports/customers', methods=['GET'])
@login_required
@admin_required
def customers_report():
    start_date = request.args.get('start_date', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    limit = request.args.get('limit', 10, type=int)

    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    sales = Sale.query.filter(Sale.sale_date.between(start_date, end_date)).all()
    customer_stats = defaultdict(lambda: {'total_amount': 0, 'count': 0})
    for sale in sales:
        key = sale.customer.name if sale.customer else 'عميل نقدي'
        bucket = customer_stats[key]
        bucket['total_amount'] += sale.total_amount
        bucket['count'] += 1

    results = sorted(customer_stats.items(), key=lambda item: item[1]['total_amount'], reverse=True)[:limit]

    return render_template('reports/customers_report.html',
                          results=results,
                          start_date=start_date,
                          end_date=end_date,
                          limit=limit)


@app.route('/activity')
@login_required
@report_required
def activity():
    logs_page = InventoryTransaction.query.options(joinedload(InventoryTransaction.user), joinedload(InventoryTransaction.product)).order_by(InventoryTransaction.timestamp.desc()).paginate(
        page=request.args.get('page', 1, type=int),
        per_page=50,
        error_out=False
    )
    type_labels = {
        'sale': 'بيع',
        'purchase': 'شراء',
        'adjustment': 'تعديل',
        'return': 'مرتجع'
    }
    return render_template('activity.html', logs=logs_page.items, logs_page=logs_page, type_labels=type_labels)


@app.route('/audit')
@login_required
@roles_required('admin', 'auditor', 'warehouse_manager')
def audit():
    audit_page = AuditLog.query.options(joinedload(AuditLog.user)).order_by(AuditLog.created_at.desc()).paginate(
        page=request.args.get('page', 1, type=int),
        per_page=50,
        error_out=False
    )
    return render_template('audit.html', audit_page=audit_page)


# Users routes
@app.route('/users')
@login_required
@permission_required('users.manage')
def users():
    # Only admin users should be able to manage users
    if not has_permission('users.manage'):
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    
    users = User.query.options(joinedload(User.branch)).order_by(User.created_at.desc()).all()
    return render_template('users/index.html', users=users)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    if not has_permission('users.manage'):
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    form = UserForm()
    populate_user_form_choices(form)
    if form.validate_on_submit():
        if User.query.filter_by(email=form.email.data).first():
            flash('البريد الإلكتروني مستخدم بالفعل من قبل مستخدم آخر.', 'danger')
            return render_template('users/add.html', form=form)
        if User.query.filter_by(username=form.username.data).first():
            flash('اسم المستخدم مستخدم بالفعل من قبل مستخدم آخر.', 'danger')
            return render_template('users/add.html', form=form)
        user = User(
            username=form.username.data,
            email=form.email.data,
            role=form.role.data,
            branch_id=form.branch_id.data if form.branch_id.data else None,
            custom_permissions=json.dumps([p for p in form.permissions.data if p in ALL_PERMISSION_CODES], ensure_ascii=False) if form.permissions.data else None,
            is_active=form.is_active.data if hasattr(form, 'is_active') else True
        )
        if not form.password.data:
            flash('يرجى تحديد كلمة مرور للمستخدم.', 'danger')
            return render_template('users/add.html', form=form)
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.flush()
        audit_log('create_user', 'User', user.id, f'إنشاء مستخدم {user.username} بصلاحية {user.role}')
        try:
            commit_with_retry()
        except OperationalError as exc:
            db.session.rollback()
            if is_database_locked_error(exc):
                flash('قاعدة البيانات مشغولة الآن. أغلق أي نسخة أخرى من البرنامج وحاول مرة أخرى.', 'danger')
                return render_template('users/add.html', form=form)
            raise
        flash('تم إضافة المستخدم بنجاح', 'success')
        return redirect(url_for('users'))
    return render_template('users/add.html', form=form)

@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    if not has_permission('users.manage'):
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    form = UserForm(obj=user)
    populate_user_form_choices(form)
    if request.method == 'GET':
        form.permissions.data = sorted(parse_permissions(user.custom_permissions))
    if form.validate_on_submit():
        user.username = form.username.data
        user.email = form.email.data
        # فقط الأدمن يمكنه تغيير الصلاحية
        if has_permission('users.manage'):
            user.role = form.role.data
            user.branch_id = form.branch_id.data if form.branch_id.data else None
            user.custom_permissions = json.dumps([p for p in form.permissions.data if p in ALL_PERMISSION_CODES], ensure_ascii=False) if form.permissions.data else None
        user.is_active = form.is_active.data
        if form.password.data:
            user.set_password(form.password.data)
        audit_log('update_user', 'User', user.id, f'تعديل مستخدم {user.username} بصلاحية {user.role}')
        try:
            commit_with_retry()
        except OperationalError as exc:
            db.session.rollback()
            if is_database_locked_error(exc):
                flash('قاعدة البيانات مشغولة الآن. أغلق أي نسخة أخرى من البرنامج وحاول مرة أخرى.', 'danger')
                return render_template('users/edit.html', form=form, user=user)
            raise
        flash('تم تحديث المستخدم بنجاح', 'success')
        return redirect(url_for('users'))
    return render_template('users/edit.html', form=form, user=user)

@app.route('/users/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    if not has_permission('users.manage'):
        flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
        return redirect(url_for('index'))
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('لا يمكنك حذف حسابك الحالي', 'danger')
        return redirect(url_for('users'))
    username = user.username
    audit_log('delete_user', 'User', user.id, f'حذف مستخدم {username}')
    db.session.delete(user)
    try:
        commit_with_retry()
    except OperationalError as exc:
        db.session.rollback()
        if is_database_locked_error(exc):
            flash('قاعدة البيانات مشغولة الآن. أغلق أي نسخة أخرى من البرنامج وحاول مرة أخرى.', 'danger')
            return redirect(url_for('users'))
        raise
    flash('تم حذف المستخدم بنجاح', 'success')
    return redirect(url_for('users'))

# Settings routes
@app.route('/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    if not has_permission('settings.manage'):
        flash('Access denied for non-admin users.', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        store_name = get_or_create_setting('store_name', 'Makhzan Store')
        store_address = get_or_create_setting('store_address', '')
        store_phone = get_or_create_setting('store_phone', '')
        store_email = get_or_create_setting('store_email', '')
        tax_rate = get_or_create_setting('tax_rate', '0')
        currency = get_or_create_setting('currency', 'SAR')
        notification_sender = get_or_create_setting('notification_sender_email', '')
        notification_password = get_or_create_setting('notification_password', '')
        notification_sender_name = get_or_create_setting('notification_sender_name', '')
        store_name.value = request.form.get('store_name', '')
        store_address.value = request.form.get('store_address', '')
        store_phone.value = request.form.get('store_phone', '')
        store_email.value = request.form.get('store_email', '')
        tax_rate.value = request.form.get('tax_rate', '0')
        currency.value = request.form.get('currency', 'SAR')
        notification_sender.value = request.form.get('notification_sender_email', '')
        notification_sender_name.value = request.form.get('notification_sender_name', '')
        password = request.form.get('notification_password')
        if password:
            notification_password.value = password
        commit_with_retry()
        flash('Settings saved successfully.', 'success')
        return redirect(url_for('settings'))

    existing_settings = {setting.key: setting.value for setting in Setting.query.all()}
    return render_template('settings/index.html',
                          store_name=existing_settings.get('store_name', 'Makhzan Store'),
                          store_address=existing_settings.get('store_address', ''),
                          store_phone=existing_settings.get('store_phone', ''),
                          store_email=existing_settings.get('store_email', ''),
                          tax_rate=existing_settings.get('tax_rate', '0'),
                          currency=existing_settings.get('currency', 'SAR'),
                          notification_sender_email=existing_settings.get('notification_sender_email', ''),
                          notification_sender_name=existing_settings.get('notification_sender_name', ''))


@app.route('/catalog/import', methods=['GET', 'POST'])
@login_required
@admin_required
def catalog_import():
    form = CatalogImportForm()
    reference_products = ReferenceProduct.query.order_by(ReferenceProduct.created_at.desc()).limit(100).all()

    if form.validate_on_submit():
        upload = form.file.data
        os.makedirs(app.instance_path, exist_ok=True)
        upload_dir = os.path.join(app.instance_path, 'reference_uploads')
        os.makedirs(upload_dir, exist_ok=True)
        filename = secure_filename(upload.filename)
        destination = os.path.join(upload_dir, filename)
        upload.save(destination)
        try:
            processed, created, updated = import_reference_catalog(destination, filename)
            flash(f'تم استيراد {processed} منتج من الملف ({created} جديد، {updated} معدل).', 'success')
        except Exception as exc:
            flash(str(exc), 'danger')
        return redirect(url_for('catalog_import'))

    return render_template('settings/catalog_import.html', form=form, reference_products=reference_products)


@app.route('/catalog/import/load-sample', methods=['POST'])
@login_required
@admin_required
def catalog_import_sample():
    flash('استيراد الملف النموذجي متوقف حتى لا يتم إدخال بيانات غير موجودة في ملفك.', 'warning')
    return redirect(url_for('catalog_import'))


@app.route('/catalog/categories-import', methods=['GET', 'POST'])
@login_required
@admin_required
def category_catalog_import():
    form = CategoryCatalogImportForm()
    reference_categories = ReferenceCategory.query.order_by(ReferenceCategory.created_at.desc()).limit(100).all()

    if form.validate_on_submit():
        upload = form.file.data
        os.makedirs(app.instance_path, exist_ok=True)
        upload_dir = os.path.join(app.instance_path, 'reference_uploads')
        os.makedirs(upload_dir, exist_ok=True)
        filename = secure_filename(upload.filename)
        destination = os.path.join(upload_dir, filename)
        upload.save(destination)
        try:
            processed, created, updated = import_reference_categories(destination, filename)
            flash(f'تم استيراد {processed} تصنيف من الملف ({created} جديد، {updated} معدل).', 'success')
        except Exception as exc:
            flash(str(exc), 'danger')
        return redirect(url_for('category_catalog_import'))

    return render_template('settings/category_import.html', form=form, reference_categories=reference_categories)


@app.route('/catalog/categories-import/load-sample', methods=['POST'])
@login_required
@admin_required
def category_catalog_import_sample():
    flash('استيراد الملف النموذجي متوقف حتى لا يتم إدخال بيانات غير موجودة في ملفك.', 'warning')
    return redirect(url_for('category_catalog_import'))


@app.route('/reference/category-suggestions')
@login_required
@admin_required
def reference_category_suggestions():
    term = (request.args.get('term') or '').strip()
    if not term:
        return jsonify([])
    limit = request.args.get('limit', 10, type=int)
    suggestions = ReferenceCategory.query.filter(
        ReferenceCategory.name.ilike(f'%{term}%')
    ).order_by(ReferenceCategory.created_at.desc()).limit(limit).all()
    return jsonify([
        {'name': ref.name, 'notes': ref.notes or ''}
        for ref in suggestions
    ])


@app.route('/reference/product-info')
@login_required
@admin_required
def reference_product_info():
    code = (request.args.get('code') or '').strip()
    if not code:
      return jsonify({'error': 'يجب إدخال الكود.'}), 400
    reference = ReferenceProduct.query.filter(
        (ReferenceProduct.barcode == code) | (ReferenceProduct.sku == code)
    ).order_by(ReferenceProduct.created_at.desc()).first()
    if not reference:
        return jsonify({'error': 'المنتج غير موجود في الكتالوج المرجعي.'}), 404
    return jsonify({
        'name': reference.name,
        'sku': reference.sku or '',
        'barcode': reference.barcode or '',
        'description': reference.description or '',
        'notes': reference.notes or '',
        'source_file': reference.source_file or ''
    })
    return jsonify({
          'name': reference.name,
          'sku': reference.sku or '',
          'barcode': reference.barcode or '',
          'description': reference.description or '',
          'notes': reference.notes or '',
          'source_file': reference.source_file or ''
      })

@app.route('/settings/theme', methods=['POST'])
@login_required
@admin_required
def set_theme():
    import json
    data = request.get_json()
    theme = data.get('theme', 'light')
    from models import Setting, db
    theme_setting = Setting.query.filter_by(key='theme').first()
    if not theme_setting:
        theme_setting = Setting(key='theme', value=theme)
        db.session.add(theme_setting)
    else:
        theme_setting.value = theme
    commit_with_retry()
    return '', 204

# Profile route
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    form = RegisterForm(obj=current_user)
    
    # Don't require password on edit
    form.password.validators = []
    form.password.flags.required = False
    form.confirm_password.validators = []
    form.confirm_password.flags.required = False
    
    # Remove role field for non-admin users
    if current_user.role != 'admin' and hasattr(form, 'role'):
        del form.role
    
    if form.validate_on_submit():
        current_user.username = form.username.data
        current_user.email = form.email.data
        current_user.full_name = form.full_name.data
        
        # Only admin can change role
        if current_user.role == 'admin' and hasattr(form, 'role'):
            current_user.role = form.role.data
        
        # Only update password if provided
        if form.password.data:
            current_user.set_password(form.password.data)
        
        db.session.commit()
        flash('تم تحديث الملف الشخصي بنجاح', 'success')
        return redirect(url_for('profile'))
    
    return render_template('profile.html', form=form)

@app.route('/print/invoice/<int:id>')
@login_required
def print_invoice(id):
    sale = Sale.query.get_or_404(id)
    from datetime import datetime
    return render_template('print/invoice.html', sale=sale, now=datetime.now())

# Due invoices and reminders
@app.route('/sales/due')
@login_required
@admin_required
def sales_due():
    due_sales = Sale.query.filter(
        Sale.payment_method == 'majel',
        Sale.status != 'cancelled',
        Sale.amount_paid < Sale.total_amount
    ).order_by(Sale.due_date.asc()).all()
    schedule_due_reminders(due_sales, 'فاتورة عميل')
    return render_template('sales/due.html', due_sales=due_sales)

@app.route('/sales/due/notify/<int:id>', methods=['POST'])
@login_required
@admin_required
def notify_due_sale(id):
    sale = Sale.query.get_or_404(id)
    if sale.payment_method != 'majel' or sale.due_amount <= 0 or sale.status == 'cancelled':
        flash('لا يمكن إرسال إشعار لفاتورة غير مستحقة أو ملغاة.', 'warning')
        return redirect(url_for('sales_due'))
    recipient = sale.contact_email
    if not recipient:
        flash('لا يوجد بريد إلكتروني لإرسال الإشعار.', 'warning')
        return redirect(url_for('sales_due'))
    days_until_due = (sale.due_date - datetime.now().date()).days if sale.due_date else None
    subject = f'تذكير بدفع فاتورة بيع {sale.invoice_number}'
    body = build_due_email_body(sale, 'فاتورة عميل', days_until_due)
    if send_due_notification_email(recipient, subject, body):
        sale.due_reminder_sent = True
        db.session.commit()
        flash('تم إرسال إشعار الدفع عبر البريد الإلكتروني.', 'success')
    else:
        flash('فشل إرسال الإشعار؛ تحقق من إعدادات البريد.', 'danger')
    return redirect(url_for('sales_due'))

@app.route('/sales/due/mark-paid/<int:id>', methods=['POST'])
@login_required
@admin_required
def mark_sale_paid(id):
    sale = Sale.query.get_or_404(id)
    sale.amount_paid = sale.total_amount
    sale.due_date = None
    sale.notification_email = None
    sale.due_reminder_sent = False
    db.session.commit()
    flash('تم تحديث الفاتورة على أنها مدفوعة.', 'success')
    return redirect(url_for('sales_due'))

@app.route('/purchases/due')
@login_required
@warehouse_required
def purchases_due():
    due_purchases = Purchase.query.filter(
        Purchase.payment_method == 'majel',
        Purchase.status != 'cancelled',
        Purchase.amount_paid < Purchase.total_amount
    ).order_by(Purchase.due_date.asc()).all()
    schedule_due_reminders(due_purchases, 'فاتورة مورد')
    completed_purchases_total = sum(purchase.total_amount for purchase in Purchase.query.filter(Purchase.status == 'completed').all())
    net_purchases_after_due = completed_purchases_total - sum(purchase.due_amount for purchase in due_purchases)
    return render_template('purchases/due.html', due_purchases=due_purchases)

@app.route('/purchases/due/notify/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def notify_due_purchase(id):
    purchase = Purchase.query.get_or_404(id)
    if purchase.payment_method != 'majel' or purchase.due_amount <= 0 or purchase.status == 'cancelled':
        flash('لا يمكن إرسال إشعار لفاتورة غير مستحقة أو ملغاة.', 'warning')
        return redirect(url_for('purchases_due'))
    recipient = purchase.contact_email
    if not recipient:
        flash('لا يوجد بريد إلكتروني لإرسال الإشعار.', 'warning')
        return redirect(url_for('purchases_due'))
    days_until_due = (purchase.due_date - datetime.now().date()).days if purchase.due_date else None
    subject = f'تذكير بدفع فاتورة شراء {purchase.invoice_number}'
    body = build_due_email_body(purchase, 'فاتورة مورد', days_until_due)
    if send_due_notification_email(recipient, subject, body):
        purchase.due_reminder_sent = True
        db.session.commit()
        flash('تم إرسال إشعار السداد عبر البريد الإلكتروني.', 'success')
    else:
        flash('فشل إرسال الإشعار؛ تحقق من إعدادات البريد.', 'danger')
    return redirect(url_for('purchases_due'))

@app.route('/purchases/due/mark-paid/<int:id>', methods=['POST'])
@login_required
@warehouse_required
def mark_purchase_paid(id):
    purchase = Purchase.query.get_or_404(id)
    purchase.amount_paid = purchase.total_amount
    purchase.due_date = None
    purchase.notification_email = None
    purchase.due_reminder_sent = False
    db.session.commit()
    flash('تم تحديث الفاتورة على أنها مدفوعة.', 'success')
    return redirect(url_for('purchases_due'))

def safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default

def parse_due_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None

def extract_due_details(total_amount, form_data):
    payment_method = form_data.get('payment_method', 'cash')
    amount_paid = safe_float(form_data.get('amount_paid'))
    due_date = parse_due_date(form_data.get('due_date'))
    notification_email = (form_data.get('notification_email') or '').strip()
    if payment_method != 'majel':
        return total_amount, None, None
    amount_paid = max(0.0, min(total_amount, amount_paid))
    return amount_paid, due_date, notification_email or None


def clean_excel_value(value):
    if value is None:
        return ''
    return str(value).strip()


def safe_excel_float(value, default=0.0):
    if value is None or value == '':
        return default
    if isinstance(value, str):
        arabic_digits = str.maketrans('٠١٢٣٤٥٦٧٨٩٫٬', '0123456789.,')
        cleaned = value.translate(arabic_digits).strip()
        numeric_chars = []
        started = False
        for char in cleaned:
            if char.isdigit():
                numeric_chars.append(char)
                started = True
            elif char in ('.', ',') and started:
                numeric_chars.append(char)
            elif started:
                break
            else:
                continue
        numeric_value = ''.join(numeric_chars).strip('.,')
        if not numeric_value:
            return default
        value = numeric_value
        last_dot = numeric_value.rfind('.')
        last_comma = numeric_value.rfind(',')
        decimal_pos = max(last_dot, last_comma)
        if last_dot != -1 and last_comma != -1:
            value = ''.join(ch for i, ch in enumerate(numeric_value) if ch.isdigit() or i == decimal_pos)
            value = value.replace(',', '.')
        elif decimal_pos != -1:
            separator = numeric_value[decimal_pos]
            digits_after = len(numeric_value) - decimal_pos - 1
            digits_before = decimal_pos
            if digits_after == 3 and digits_before > 3:
                value = numeric_value.replace(separator, '')
            else:
                value = numeric_value.replace(separator, '.')
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def safe_excel_int(value, default=0):
    if value is None or value == '':
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def find_product_for_import(name, sku='', barcode=''):
    product = None
    if barcode:
        product = Product.query.filter_by(barcode=barcode).first()
    if not product and sku:
        product = Product.query.filter_by(sku=sku).first()
    if not product:
        product = Product.query.filter(func.lower(Product.name) == name.lower()).first()
    return product


def find_or_create_category_for_import(name, notes=''):
    if not name:
        return None, False

    category = Category.query.filter(func.lower(Category.name) == name.lower()).first()
    if category:
        if notes:
            if not category.description:
                category.description = notes
            category.notes = notes
        return category, False

    category = Category(name=name, description=notes or None, notes=notes or None)
    db.session.add(category)
    db.session.flush()
    return category, True


def reset_product_catalog_for_import():
    if SaleItem.query.first() or PurchaseItem.query.first():
        raise ValueError('لا يمكن استبدال الكتالوج لأن هناك فواتير بيع أو شراء مرتبطة بالمنتجات الحالية.')

    InventoryTransaction.query.delete(synchronize_session=False)
    Inventory.query.delete(synchronize_session=False)
    Product.query.delete(synchronize_session=False)
    ReferenceProduct.query.delete(synchronize_session=False)
    Category.query.delete(synchronize_session=False)
    ReferenceCategory.query.delete(synchronize_session=False)
    db.session.flush()


def import_reference_catalog(filepath, source_label=None, replace_existing=False):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('مطلوب تثبيت مكتبة openpyxl لقراءة ملفات Excel.') from exc

    # ملف مش موجود
    if not os.path.exists(filepath):
        raise ValueError('مسار ملف Excel غير موجود.')

    workbook = load_workbook(filepath, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    header_row = next(rows, None)
    if not header_row:
        raise ValueError('ملف Excel لا يحتوي على صف عناوين (Header).')

    # تجهيز العناوين
    normalized_headers = [str(col).strip().lower() if col else '' for col in header_row]

    def find_index(keywords, exclude=None, exact_keywords=None):
        exclude = set(exclude or [])
        exact_keywords = exact_keywords or []
        for idx, header in enumerate(normalized_headers):
            if idx in exclude:
                continue
            if any(header == keyword for keyword in exact_keywords):
                return idx
        for idx, header in enumerate(normalized_headers):
            if idx in exclude:
                continue
            if any(keyword in header for keyword in keywords):
                return idx
        return None

    # البحث عن الأعمدة
    name_idx = find_index(['اسم المنتج', 'product', 'product name'], exact_keywords=['اسم المنتج'])
    category_idx = find_index(['تصنيف', 'category', 'classification'], exact_keywords=['التصنيف'])
    barcode_idx = find_index(['barcode', 'code', 'باركود'], exact_keywords=['الباركود', 'barcode'])
    sku_idx = find_index(['sku', 'رمز المنتج'], exact_keywords=['رمز المنتج (sku)', 'sku'])
    purchase_price_idx = find_index(['purchase price', 'cost', 'buy price', 'سعر الشراء'], exact_keywords=['سعر الشراء'])
    sale_price_idx = find_index(['sale price', 'selling price', 'سعر البيع'], exact_keywords=['سعر البيع'])
    min_quantity_idx = find_index(['min quantity', 'minimum quantity', 'min stock', 'الحد الأدنى'], exact_keywords=['الحد الأدنى للكمية'])
    quantity_idx = find_index(
        ['initial quantity', 'opening quantity', 'start quantity', 'الكمية الأولية'],
        exclude=[min_quantity_idx],
        exact_keywords=['الكمية الأولية']
    )
    if quantity_idx is None:
        quantity_idx = find_index(['quantity', 'qty', 'الكمية'], exclude=[min_quantity_idx])
    desc_idx = find_index(['desc', 'description', 'وصف'], exact_keywords=['الوصف'])
    notes_idx = find_index(['notes', 'ملاحظات', 'remarks', 'comments'])
    if sale_price_idx is None:
        sale_price_idx = find_index(['price'], exclude=[purchase_price_idx])
        if sale_price_idx == purchase_price_idx:
            sale_price_idx = None

    # لو مفيش عمود اسم المنتج
    if name_idx is None:
        raise ValueError('لم يتم العثور على عمود اسم المنتج داخل ملف Excel.')

    if replace_existing:
        reset_product_catalog_for_import()

    processed = created = updated = 0
    source_label = source_label or os.path.basename(filepath)

    for row in rows:
        if not row:
            continue

        name_cell = row[name_idx] if name_idx < len(row) else None
        if not name_cell:
            continue

        name = str(name_cell).strip()
        if not name:
            continue

        sku = clean_excel_value(row[sku_idx]) if sku_idx is not None and sku_idx < len(row) else ''
        barcode = clean_excel_value(row[barcode_idx]) if barcode_idx is not None and barcode_idx < len(row) else ''
        description = clean_excel_value(row[desc_idx]) if desc_idx is not None and desc_idx < len(row) else ''
        notes = clean_excel_value(row[notes_idx]) if notes_idx is not None and notes_idx < len(row) else ''
        category_name = clean_excel_value(row[category_idx]) if category_idx is not None and category_idx < len(row) else ''
        purchase_price = safe_excel_float(row[purchase_price_idx]) if purchase_price_idx is not None and purchase_price_idx < len(row) else 0.0
        sale_price = safe_excel_float(row[sale_price_idx]) if sale_price_idx is not None and sale_price_idx < len(row) else 0.0
        min_quantity = safe_excel_int(row[min_quantity_idx]) if min_quantity_idx is not None and min_quantity_idx < len(row) else 0
        quantity = safe_excel_int(row[quantity_idx]) if quantity_idx is not None and quantity_idx < len(row) else 0

        processed += 1

        reference = None

        # البحث لو موجود قبل كده
        if barcode:
            reference = ReferenceProduct.query.filter_by(barcode=barcode).first()

        if not reference and sku:
            reference = ReferenceProduct.query.filter_by(sku=sku).first()

        if not reference:
            normalized_name = name.lower()
            reference = ReferenceProduct.query.filter(func.lower(ReferenceProduct.name) == normalized_name).first()

        # إنشاء جديد
        if not reference:
            reference = ReferenceProduct(
                name=name,
                sku=sku or None,
                barcode=barcode or None,
                description=description or None,
                notes=notes or None,
                source_file=source_label
            )
            db.session.add(reference)

        # تحديث القديم
        else:
            reference.name = name
            reference.sku = sku or reference.sku
            reference.barcode = barcode or reference.barcode
            reference.description = description or reference.description
            reference.notes = notes or reference.notes
            reference.source_file = source_label

        category = None
        if category_name:
            category, _ = find_or_create_category_for_import(category_name)

        product = find_product_for_import(name, sku, barcode)
        if not product:
            product = Product(
                name=name,
                description=description or notes or None,
                barcode=barcode or None,
                sku=sku or None,
                purchase_price=purchase_price,
                sale_price=sale_price,
                min_quantity=min_quantity,
                category_id=category.id if category else None
            )
            db.session.add(product)
            db.session.flush()
            db.session.add(Inventory(product_id=product.id, quantity=quantity))
            created += 1
        else:
            product.name = name
            if description or notes:
                product.description = description or notes
            if sku and not product.sku:
                product.sku = sku
            if barcode and not product.barcode:
                product.barcode = barcode
            if purchase_price_idx is not None:
                product.purchase_price = purchase_price
            if sale_price_idx is not None:
                product.sale_price = sale_price
            if min_quantity_idx is not None:
                product.min_quantity = min_quantity
            if category and not product.category_id:
                product.category_id = category.id
            if not product.inventory:
                db.session.add(Inventory(product_id=product.id, quantity=quantity))
            updated += 1

    if processed:
        db.session.commit()

    return processed, created, updated

def reset_category_catalog_for_import():
    ReferenceCategory.query.delete(synchronize_session=False)
    if not Product.query.first():
        Category.query.delete(synchronize_session=False)
    db.session.flush()


def import_reference_categories(filepath, source_label=None, replace_existing=False):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is required to read Excel files.') from exc

    if not os.path.exists(filepath):
        raise ValueError('ملف تصنيفات المراجع غير موجود.')

    workbook = load_workbook(filepath, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError('ملف Excel لا يحتوي على صف رؤوس.')

    normalized_headers = [str(col).strip().lower() if col else '' for col in header_row]

    def find_index(keywords):
        for idx, header in enumerate(normalized_headers):
            if any(keyword in header for keyword in keywords):
                return idx
        return None

    name_idx = find_index(['تصنيف', 'classification', 'category', 'name'])
    notes_idx = find_index(['notes', 'ملاحظات', 'remarks', 'comments'])

    if name_idx is None:
        raise ValueError('الملف لا يحتوي على عمود اسم التصنيف.')

    if replace_existing:
        reset_category_catalog_for_import()

    processed = created = updated = 0
    source_label = source_label or os.path.basename(filepath)

    for row in rows:
        if not row:
            continue
        name_cell = row[name_idx] if name_idx < len(row) else None
        if not name_cell:
            continue
        name = str(name_cell).strip()
        if not name:
            continue
        notes = ''
        if notes_idx is not None and notes_idx < len(row) and row[notes_idx]:
            notes = str(row[notes_idx]).strip()

        processed += 1
        normalized_name = name.lower()
        reference = ReferenceCategory.query.filter(func.lower(ReferenceCategory.name) == normalized_name).first()

        if not reference:
            reference = ReferenceCategory(
                name=name,
                notes=notes or None,
                source_file=source_label
            )
            db.session.add(reference)
        else:
            reference.name = name
            reference.notes = notes or reference.notes
            reference.source_file = source_label

        _, was_created = find_or_create_category_for_import(name, notes)
        if was_created:
            created += 1
        else:
            updated += 1

    if processed:
        db.session.commit()

    return processed, created, updated


def get_notification_config():
    keys = ['notification_sender_email', 'notification_password', 'notification_sender_name']
    settings = Setting.query.filter(Setting.key.in_(keys)).all()
    return {s.key: s.value for s in settings}


def send_due_notification_email(recipient, subject, body):
    config = get_notification_config()
    sender = config.get('notification_sender_email')
    password = config.get('notification_password')
    sender_name = config.get('notification_sender_name') or sender
    if not sender or not password or not recipient:
        return False
    msg = MIMEText(body, 'html')
    msg['Subject'] = subject
    msg['From'] = f"{sender_name} <{sender}>"
    msg['To'] = recipient
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(sender, password)
            smtp.send_message(msg)
        return True
    except Exception as exc:
        app.logger.error('Failed to send reminder email: %s', exc)
        return False


def build_due_email_body(record, entity_label, days_until_due=None):
    due_date = record.due_date.strftime('%Y-%m-%d') if record.due_date else 'غير محدد'
    header = f'<p>تذكير بالفاتورة {entity_label} <strong>{record.invoice_number}</strong>.</p>'
    totals = f'<p>الإجمالي: {record.total_amount:.2f} &nbsp; مدفوع: {record.amount_paid:.2f} &nbsp; المتبقي: {record.due_amount:.2f}.</p>'
    date_line = f'<p>تاريخ الاستحقاق: {due_date}.</p>'
    reminder_line = f'<p>باقي {days_until_due} يوم.</p>' if days_until_due is not None else ''
    preamble = '<p>يرجى مراجعة التذكير قبل تاريخ الاستحقاق.</p>'
    closing = '<p>برجاء تسوية الرصيد المتبقي.</p>'
    return f"{preamble}{header}{totals}{date_line}{reminder_line}{closing}"


def schedule_due_reminders(records, entity_label):
    today = datetime.now().date()
    updated = False
    for record in records:
        if record.status == 'cancelled' or record.payment_method != 'majel':
            continue
        if not record.due_date or record.due_amount <= 0:
            continue
        if record.due_reminder_sent:
            continue
        days_until_due = (record.due_date - today).days
        if days_until_due < 0 or days_until_due > 2:
            continue
        recipient = record.contact_email
        if not recipient:
            continue
        subject = f'تذكير استحقاق {entity_label} {record.invoice_number}'
        body = build_due_email_body(record, entity_label, days_until_due)
        if send_due_notification_email(recipient, subject, body):
            record.due_reminder_sent = True
            updated = True
    if updated:
        db.session.commit()


# Add more routes for purchases, sales, inventory, etc.

@app.context_processor
def inject_settings():
    from models import Setting
    settings = {s.key: s.value for s in Setting.query.all()}

    def page_url(page):
        args = request.args.to_dict(flat=True)
        args['page'] = page
        return url_for(request.endpoint, **(request.view_args or {}), **args)

    return dict(settings=settings, page_url=page_url, can=has_permission, permission_groups=PERMISSION_GROUPS)

if __name__ == '__main__':
    app.run(debug=True, port=5001, use_reloader=False)
